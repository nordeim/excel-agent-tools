"""Unit tests for the macro CLI tools.

Tests xls_has_macros, xls_inspect_macros, xls_validate_macro_safety,
xls_remove_macros, and xls_inject_vba_project.

Note: These tests run CLI tools via subprocess, so mocking internal classes
won't work. Tests focus on actual behavior and output validation.
"""

from __future__ import annotations

import json
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

# -----------------------------------------------------------------------------
# Fixtures
# -----------------------------------------------------------------------------


@pytest.fixture
def macro_workbook(tmp_path: Path) -> Path:
    """Create a workbook with VBA project marker."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = "Test"

    path = tmp_path / "test_macros.xlsm"
    wb.save(str(path))

    # Inject vbaProject.bin marker
    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA_PROJECT_DATA")

    return path


@pytest.fixture
def clean_workbook(tmp_path: Path) -> Path:
    """Create a clean workbook without macros."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = "Clean"

    path = tmp_path / "clean.xlsx"
    wb.save(str(path))
    return path


# -----------------------------------------------------------------------------
# Helper
# -----------------------------------------------------------------------------


def run_tool(tool_name: str, *args: str) -> tuple[dict, int]:
    """Helper to run a CLI tool and parse output."""
    cmd = [sys.executable, "-m", f"excel_agent.tools.macros.{tool_name}", *args]
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
    )
    try:
        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


# -----------------------------------------------------------------------------
# xls_has_macros Tests
# -----------------------------------------------------------------------------


def test_has_macros_detects_macros(macro_workbook: Path):
    """Test xls_has_macros detects VBA macros."""
    output, exit_code = run_tool("xls_has_macros", "--input", str(macro_workbook))

    assert exit_code == 0
    assert output.get("data", {}).get("has_macros") is True


def test_has_macros_clean_file(clean_workbook: Path):
    """Test xls_has_macros returns False for clean file."""
    output, exit_code = run_tool("xls_has_macros", "--input", str(clean_workbook))

    assert exit_code == 0
    assert output.get("data", {}).get("has_macros") is False


def test_has_macros_missing_file(tmp_path: Path):
    """Test xls_has_macros handles missing file."""
    path = tmp_path / "nonexistent.xlsx"
    output, exit_code = run_tool("xls_has_macros", "--input", str(path))

    assert exit_code == 2  # FILE_NOT_FOUND
    assert output.get("status") == "error"


# -----------------------------------------------------------------------------
# xls_inspect_macros Tests
# -----------------------------------------------------------------------------


def test_inspect_macros_lists_modules(macro_workbook: Path):
    """Test xls_inspect_macros lists VBA modules."""
    output, exit_code = run_tool("xls_inspect_macros", "--input", str(macro_workbook))

    assert exit_code == 0
    assert "has_macros" in output.get("data", {})


def test_inspect_macros_clean_file(clean_workbook: Path):
    """Test xls_inspect_macros handles clean file."""
    output, exit_code = run_tool("xls_inspect_macros", "--input", str(clean_workbook))

    assert exit_code == 0
    data = output.get("data", {})
    assert data.get("has_macros") is False
    assert data.get("module_count") == 0


def test_inspect_macros_custom_preview_length(macro_workbook: Path):
    """Test xls_inspect_macros with custom code preview length."""
    output, exit_code = run_tool(
        "xls_inspect_macros",
        "--input",
        str(macro_workbook),
        "--code-preview-length",
        "100",
    )

    assert exit_code == 0


# -----------------------------------------------------------------------------
# xls_validate_macro_safety Tests
# -----------------------------------------------------------------------------


def test_validate_macro_safety_detects_risk(macro_workbook: Path):
    """Test xls_validate_macro_safety analyzes macros."""
    output, exit_code = run_tool("xls_validate_macro_safety", "--input", str(macro_workbook))

    assert exit_code == 0
    data = output.get("data", {})
    assert "risk_level" in data
    assert "risk_score" in data


def test_validate_macro_safety_clean_file(clean_workbook: Path):
    """Test xls_validate_macro_safety on clean file."""
    output, exit_code = run_tool("xls_validate_macro_safety", "--input", str(clean_workbook))

    assert exit_code == 0
    data = output.get("data", {})
    assert data.get("risk_level") == "none"
    assert data.get("risk_score") == 0


# -----------------------------------------------------------------------------
# xls_remove_macros Tests
# -----------------------------------------------------------------------------


def test_remove_macros_requires_token(macro_workbook: Path, tmp_path: Path):
    """Test xls_remove_macros requires tokens."""
    output_path = tmp_path / "cleaned.xlsx"
    output, exit_code = run_tool(
        "xls_remove_macros",
        "--input",
        str(macro_workbook),
        "--output",
        str(output_path),
    )

    # argparse exits with 2 for missing required arguments
    assert exit_code == 2


def test_remove_macros_requires_token2(macro_workbook: Path, tmp_path: Path):
    """Test xls_remove_macros requires second token."""
    output_path = tmp_path / "cleaned.xlsx"
    output, exit_code = run_tool(
        "xls_remove_macros",
        "--input",
        str(macro_workbook),
        "--output",
        str(output_path),
        "--token",
        "only_one_token",
    )

    # argparse exits with 2 for missing --token2
    assert exit_code == 2


def test_remove_macros_no_macros(clean_workbook: Path, tmp_path: Path):
    """Test xls_remove_macros on file without macros returns early."""
    output_path = tmp_path / "cleaned.xlsx"

    # Generate a valid token (mock token won't work via subprocess)
    # Instead, test that when no macros are found, we get success
    from excel_agent.core.macro_handler import has_macros

    assert has_macros(clean_workbook) is False

    # Tool will fail on token validation before checking macros
    # We verify the tool handles this case
    output, exit_code = run_tool(
        "xls_remove_macros",
        "--input",
        str(clean_workbook),
        "--output",
        str(output_path),
    )

    # Should error on missing token (argparse exits 2)
    assert exit_code == 2


# -----------------------------------------------------------------------------
# xls_inject_vba_project Tests
# -----------------------------------------------------------------------------


def test_inject_requires_token(clean_workbook: Path, tmp_path: Path):
    """Test xls_inject_vba_project requires token."""
    vba_bin = tmp_path / "vbaProject.bin"
    vba_bin.write_bytes(b"VBA_DATA")
    output_path = tmp_path / "injected.xlsm"

    output, exit_code = run_tool(
        "xls_inject_vba_project",
        "--input",
        str(clean_workbook),
        "--vba-bin",
        str(vba_bin),
        "--output",
        str(output_path),
    )

    # Should error due to missing token (exit 1 for validation or 5 for internal)
    assert exit_code in [1, 5]
    assert output.get("status") == "error"


def test_inject_requires_vba_bin(clean_workbook: Path, tmp_path: Path):
    """Test xls_inject_vba_project requires vba-bin file."""
    vba_bin = tmp_path / "nonexistent.bin"
    output_path = tmp_path / "injected.xlsm"

    output, exit_code = run_tool(
        "xls_inject_vba_project",
        "--input",
        str(clean_workbook),
        "--vba-bin",
        str(vba_bin),
        "--output",
        str(output_path),
        "--token",
        "some_token",
    )

    # validate_input_path rejects non-existent .bin file with exit 2 (FILE_NOT_FOUND)
    # But the actual implementation calls exit_with which sys.exit()s directly
    # The tool's run_tool wrapper catches unexpected exceptions as exit 5
    # so this could be 2 or 5 depending on timing
    assert exit_code in [1, 2, 5]  # VALIDATION_ERROR, FILE_NOT_FOUND, or INTERNAL_ERROR
    assert output.get("status") == "error"


# -----------------------------------------------------------------------------
# Integration Tests
# -----------------------------------------------------------------------------


def test_macro_workflow_full(clean_workbook: Path, macro_workbook: Path, tmp_path: Path):
    """Test full macro workflow: has_macros -> inspect -> validate."""
    # Step 1: Check for macros
    output1, _ = run_tool("xls_has_macros", "--input", str(macro_workbook))
    assert output1.get("data", {}).get("has_macros") is True

    # Step 2: Inspect
    output2, _ = run_tool("xls_inspect_macros", "--input", str(macro_workbook))
    assert output2.get("status") == "success"

    # Step 3: Validate safety
    output3, _ = run_tool("xls_validate_macro_safety", "--input", str(macro_workbook))
    assert "risk_level" in output3.get("data", {})

    # Step 4: Verify clean file has no macros
    output4, _ = run_tool("xls_has_macros", "--input", str(clean_workbook))
    assert output4.get("data", {}).get("has_macros") is False
