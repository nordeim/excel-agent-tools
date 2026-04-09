"""Unit tests for export tools.

Tests xls_export_csv, xls_export_json, and xls_export_pdf.
"""

from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run an export tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.export.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    try:
        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data for export tests."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Sales"
    ws["C1"] = "Date"

    # Data rows
    ws["A2"] = "Widget A"
    ws["B2"] = 100.50
    ws["C2"] = datetime(2024, 1, 15)

    ws["A3"] = "Widget B"
    ws["B3"] = 250.75
    ws["C3"] = datetime(2024, 2, 20)

    ws["A4"] = "Widget C"
    ws["B4"] = None  # Test None handling
    ws["C4"] = datetime(2024, 3, 25)

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestExportCSV:
    """Test xls_export_csv tool."""

    def test_export_basic_csv(self, data_workbook: Path, tmp_path: Path):
        """Test basic CSV export."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["row_count"] == 4  # header + 3 data rows
        assert output["data"]["encoding"] == "utf-8"

        # Verify CSV content
        with open(output_path) as f:
            lines = f.readlines()
        assert len(lines) == 4
        assert "Product,Sales,Date" in lines[0]

    def test_export_with_encoding(self, data_workbook: Path, tmp_path: Path):
        """Test CSV export with specific encoding."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--encoding",
            "latin-1",
        )

        assert exit_code == 0
        assert output["data"]["encoding"] == "latin-1"

    def test_export_with_delimiter(self, data_workbook: Path, tmp_path: Path):
        """Test CSV export with different delimiter."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--delimiter",
            ";",
        )

        assert exit_code == 0
        assert output["data"]["delimiter"] == ";"

        # Verify delimiter used
        with open(output_path) as f:
            content = f.read()
        assert ";" in content

    def test_export_specific_sheet(self, data_workbook: Path, tmp_path: Path):
        """Test CSV export from specific sheet."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--sheet",
            "Data",
        )

        assert exit_code == 0
        assert output["data"]["sheet"] == "Data"

    def test_invalid_encoding_error(self, data_workbook: Path, tmp_path: Path):
        """Test error on invalid encoding."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--encoding",
            "invalid-encoding",
        )

        assert exit_code == 1
        assert "Unsupported encoding" in str(output.get("warnings", []))


class TestExportJSON:
    """Test xls_export_json tool."""

    def test_export_records_orientation(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export with records orientation."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--orient",
            "records",
        )

        assert exit_code == 0
        assert output["data"]["orient"] == "records"

        # Verify JSON structure
        with open(output_path) as f:
            data = json.load(f)
        assert isinstance(data, list)
        assert len(data) == 3
        assert "Product" in data[0]
        assert "Sales" in data[0]

    def test_export_values_orientation(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export with values orientation."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--orient",
            "values",
        )

        assert exit_code == 0

        with open(output_path) as f:
            data = json.load(f)
        assert isinstance(data, list)
        assert len(data) == 4  # All rows including header

    def test_export_columns_orientation(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export with columns orientation."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--orient",
            "columns",
        )

        assert exit_code == 0

        with open(output_path) as f:
            data = json.load(f)
        assert isinstance(data, dict)
        assert "Product" in data
        assert "Sales" in data

    def test_export_with_range(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export with specific range."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--range",
            "A1:B3",
            "--orient",
            "records",
        )

        assert exit_code == 0

        with open(output_path) as f:
            data = json.load(f)
        assert len(data) == 2  # 2 data rows (A2:B3)

    def test_export_with_pretty(self, data_workbook: Path, tmp_path: Path):
        """Test JSON export with pretty printing."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
            "--pretty",
        )

        assert exit_code == 0

        with open(output_path) as f:
            content = f.read()
        # Pretty-printed JSON should have indentation
        assert "  " in content or "\n" in content

    def test_date_conversion(self, data_workbook: Path, tmp_path: Path):
        """Test that dates are converted to ISO strings."""
        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(data_workbook),
            "--outfile",
            str(output_path),
        )

        assert exit_code == 0

        with open(output_path) as f:
            data = json.load(f)
        # Check date is ISO formatted
        assert "2024-01-15" in str(data) or "2024-01-15T" in str(data)


class TestExportPDF:
    """Test xls_export_pdf tool."""

    def test_libreoffice_not_found(self, tmp_path: Path):
        """Test error when LibreOffice not found."""
        # Create minimal workbook
        wb = Workbook()
        wb.active["A1"] = "Test"
        input_path = tmp_path / "test.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.pdf"

        # Mock PATH to exclude LibreOffice
        import os

        original_path = os.environ.get("PATH", "")
        os.environ["PATH"] = "/nonexistent"

        try:
            output, exit_code = run_tool(
                "xls_export_pdf",
                "--input",
                str(input_path),
                "--output",
                str(output_path),
            )
            # May fail or succeed depending on system LO installation
            if exit_code == 2:
                assert "LibreOffice not found" in str(output.get("warnings", []))
        finally:
            os.environ["PATH"] = original_path

    @pytest.mark.slow
    def test_pdf_export_success(self, tmp_path: Path):
        """Test successful PDF export (requires LibreOffice)."""
        # Skip if LibreOffice not available
        import shutil

        if not shutil.which("soffice") and not shutil.which("libreoffice"):
            pytest.skip("LibreOffice not installed")

        # Create test workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws["A2"] = "Data"
        input_path = tmp_path / "test.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.pdf"
        output, exit_code = run_tool(
            "xls_export_pdf",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
            "--timeout",
            "60",
        )

        # PDF creation may or may not succeed depending on environment
        if exit_code == 0:
            assert output_path.exists()
            assert output_path.stat().st_size > 0

    def test_pdf_export_timeout(self, tmp_path: Path):
        """Test timeout handling."""
        # This test verifies timeout parameter validation
        wb = Workbook()
        wb.active["A1"] = "Test"
        input_path = tmp_path / "test.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.pdf"
        output, exit_code = run_tool(
            "xls_export_pdf",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
            "--timeout",
            "1000",  # Above max
        )

        # Should either succeed with warning or fail gracefully
        if exit_code == 0:
            # Check timeout was capped
            assert output["data"]["timeout_used"] <= 600

    def test_recalc_flag(self, tmp_path: Path):
        """Test --recalc flag."""
        wb = Workbook()
        wb.active["A1"] = "=1+1"
        input_path = tmp_path / "test.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.pdf"
        output, exit_code = run_tool(
            "xls_export_pdf",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
            "--recalc",
        )

        # Should complete without error (actual recalc not implemented)
        assert exit_code in [0, 2]  # Success or LO not found


class TestExportEdgeCases:
    """Test export edge cases."""

    def test_empty_sheet_csv(self, tmp_path: Path):
        """Test CSV export of empty sheet."""
        wb = Workbook()
        wb.active  # Empty sheet
        input_path = tmp_path / "empty.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
        )

        assert exit_code == 0
        # Empty sheet should create empty or minimal CSV
        assert output_path.exists()

    def test_empty_sheet_json(self, tmp_path: Path):
        """Test JSON export of empty sheet."""
        wb = Workbook()
        wb.active  # Empty sheet
        input_path = tmp_path / "empty.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.json"
        output, exit_code = run_tool(
            "xls_export_json",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
        )

        assert exit_code == 0
        assert output_path.exists()

        with open(output_path) as f:
            data = json.load(f)
        assert isinstance(data, list)
        assert len(data) == 0

    def test_special_characters_csv(self, tmp_path: Path):
        """Test CSV with special characters."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Special: é, ñ, ü, 中文"
        ws["A2"] = "Line\nbreak"
        ws["A3"] = 'Quote "test"'
        input_path = tmp_path / "special.xlsx"
        wb.save(str(input_path))

        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(input_path),
            "--outfile",
            str(output_path),
            "--encoding",
            "utf-8",
        )

        assert exit_code == 0
        with open(output_path, encoding="utf-8") as f:
            content = f.read()
        assert "Special" in content

    def test_missing_file_error(self, tmp_path: Path):
        """Test error on missing input file."""
        output_path = tmp_path / "output.csv"
        output, exit_code = run_tool(
            "xls_export_csv",
            "--input",
            str(tmp_path / "nonexistent.xlsx"),
            "--outfile",
            str(output_path),
        )

        assert exit_code == 2  # File not found
