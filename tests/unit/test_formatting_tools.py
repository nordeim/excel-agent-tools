"""Unit tests for formatting tools.

Tests xls_format_range, xls_set_column_width, xls_freeze_panes,
xls_apply_conditional_formatting, and xls_set_number_format.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run a formatting tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.formatting.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data for formatting tests."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"

    # Data rows
    for i in range(2, 6):
        ws[f"A{i}"] = f"Product {i - 1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = i * 250

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestFormatRange:
    """Test xls_format_range tool."""

    def test_format_font(self, data_workbook: Path, tmp_path: Path):
        """Test applying font formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
            "--font",
            '{"name": "Arial", "size": 14, "bold": true}',
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["font_applied"] is True

        # Verify formatting
        wb = load_workbook(str(output_path))
        ws = wb.active
        cell = ws["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14

    def test_format_fill(self, data_workbook: Path, tmp_path: Path):
        """Test applying fill formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
            "--fill",
            '{"fgColor": "FFFF00", "patternType": "solid"}',
        )

        assert exit_code == 0
        assert output["data"]["fill_applied"] is True

    def test_format_border(self, data_workbook: Path, tmp_path: Path):
        """Test applying border formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
            "--border",
            '{"top": {"style": "thin", "color": "000000"}, "bottom": {"style": "thin"}}',
        )

        assert exit_code == 0
        assert output["data"]["border_applied"] is True

    def test_format_alignment(self, data_workbook: Path, tmp_path: Path):
        """Test applying alignment formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
            "--alignment",
            '{"horizontal": "center", "vertical": "center", "wrapText": true}',
        )

        assert exit_code == 0
        assert output["data"]["alignment_applied"] is True

        # Verify alignment
        wb = load_workbook(str(output_path))
        ws = wb.active
        cell = ws["A1"]
        assert cell.alignment.horizontal == "center"

    def test_format_combined(self, data_workbook: Path, tmp_path: Path):
        """Test applying multiple formatting types."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
            "--font",
            '{"bold": true}',
            "--fill",
            '{"fgColor": "CCCCCC"}',
            "--border",
            '{"top": {"style": "thin"}}',
            "--alignment",
            '{"horizontal": "center"}',
        )

        assert exit_code == 0
        assert output["data"]["font_applied"] is True
        assert output["data"]["fill_applied"] is True
        assert output["data"]["border_applied"] is True
        assert output["data"]["alignment_applied"] is True

    def test_no_style_error(self, data_workbook: Path, tmp_path: Path):
        """Test error when no style provided."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D1",
        )

        assert exit_code == 1
        assert "At least one style" in str(output.get("warnings", []))


class TestSetColumnWidth:
    """Test xls_set_column_width tool."""

    def test_fixed_width(self, data_workbook: Path, tmp_path: Path):
        """Test setting fixed column width."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_column_width",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--columns",
            "A",
            "--width",
            "20",
        )

        assert exit_code == 0
        assert output["data"]["width"] == "20"

        # Verify width
        wb = load_workbook(str(output_path))
        ws = wb.active
        assert ws.column_dimensions["A"].width == 20.0

    def test_multiple_columns(self, data_workbook: Path, tmp_path: Path):
        """Test setting width on multiple columns."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_column_width",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--columns",
            "A,B,C",
            "--width",
            "15",
        )

        assert exit_code == 0
        assert len(output["data"]["columns_affected"]) == 3

    def test_column_range(self, data_workbook: Path, tmp_path: Path):
        """Test setting width on column range."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_column_width",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--columns",
            "A:C",
            "--width",
            "12",
        )

        assert exit_code == 0
        assert len(output["data"]["columns_affected"]) == 3

    def test_auto_fit(self, data_workbook: Path, tmp_path: Path):
        """Test auto-fit column width."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_column_width",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--columns",
            "A",
            "--width",
            "auto",
        )

        assert exit_code == 0
        assert output["data"]["is_auto"] is True

    def test_invalid_width(self, data_workbook: Path, tmp_path: Path):
        """Test error on invalid width value."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_column_width",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--columns",
            "A",
            "--width",
            "300",  # Exceeds max
        )

        assert exit_code == 1


class TestFreezePanes:
    """Test xls_freeze_panes tool."""

    def test_freeze_first_row(self, data_workbook: Path, tmp_path: Path):
        """Test freezing first row (A2)."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_freeze_panes",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--freeze",
            "A2",
        )

        assert exit_code == 0
        assert output["data"]["frozen_rows"] == 1
        assert output["data"]["frozen_cols"] == 0

    def test_freeze_first_column(self, data_workbook: Path, tmp_path: Path):
        """Test freezing first column (B1)."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_freeze_panes",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--freeze",
            "B1",
        )

        assert exit_code == 0
        assert output["data"]["frozen_rows"] == 0
        assert output["data"]["frozen_cols"] == 1

    def test_freeze_both(self, data_workbook: Path, tmp_path: Path):
        """Test freezing first row and column (B2)."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_freeze_panes",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--freeze",
            "B2",
        )

        assert exit_code == 0
        assert output["data"]["frozen_rows"] == 1
        assert output["data"]["frozen_cols"] == 1

    def test_unfreeze(self, data_workbook: Path, tmp_path: Path):
        """Test unfreezing panes."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_freeze_panes",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--freeze",
            "none",
        )

        assert exit_code == 0
        assert output["data"]["is_frozen"] is False


class TestConditionalFormatting:
    """Test xls_apply_conditional_formatting tool."""

    def test_cellis_rule(self, data_workbook: Path, tmp_path: Path):
        """Test cellIs conditional formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--type",
            "cellis",
            "--config",
            '{"operator": "greaterThan", "formula": ["200"], "fill": {"fgColor": "FF0000"}}',
        )

        assert exit_code == 0
        assert output["data"]["type"] == "cellis"

    def test_colorscale_rule(self, data_workbook: Path, tmp_path: Path):
        """Test colorScale conditional formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--type",
            "colorscale",
            "--config",
            '{"start_type": "min", "start_color": "FF0000", "end_type": "max", "end_color": "00FF00"}',
        )

        assert exit_code == 0
        assert output["data"]["type"] == "colorscale"

    def test_databar_rule(self, data_workbook: Path, tmp_path: Path):
        """Test dataBar conditional formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--type",
            "databar",
            "--config",
            '{"start_type": "min", "end_type": "max", "color": "638EC6"}',
        )

        assert exit_code == 0
        assert output["data"]["type"] == "databar"

    def test_iconset_rule(self, data_workbook: Path, tmp_path: Path):
        """Test iconSet conditional formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--type",
            "iconset",
            "--config",
            '{"icon_style": "3Arrows", "type": "percent", "values": [0, 33, 67]}',
        )

        assert exit_code == 0
        assert output["data"]["type"] == "iconset"

    def test_formula_rule(self, data_workbook: Path, tmp_path: Path):
        """Test formula conditional formatting."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--type",
            "formula",
            "--config",
            '{"formula": ["MOD(B2,2)=0"], "fill": {"fgColor": "0000FF"}}',
        )

        assert exit_code == 0
        assert output["data"]["type"] == "formula"

    def test_invalid_type(self, data_workbook: Path, tmp_path: Path):
        """Test error on invalid type."""
        output_path = tmp_path / "output.xlsx"
        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "excel_agent.tools.formatting.xls_apply_conditional_formatting",
                "--input",
                str(data_workbook),
                "--output",
                str(output_path),
                "--range",
                "A1:A10",
                "--type",
                "invalid",
                "--config",
                "{}",
            ],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 2  # argparse error


class TestNumberFormat:
    """Test xls_set_number_format tool."""

    def test_currency_format(self, data_workbook: Path, tmp_path: Path):
        """Test currency number format."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--number-format",
            '"$"#,##0.00',
        )

        assert exit_code == 0
        assert output["data"]["format"] == '"$"#,##0.00'

        # Verify format
        wb = load_workbook(str(output_path))
        ws = wb.active
        cell = ws["B2"]
        assert cell.number_format == '"$"#,##0.00'

    def test_percentage_format(self, data_workbook: Path, tmp_path: Path):
        """Test percentage number format."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "B2:D5",
            "--number-format",
            "0.00%",
        )

        assert exit_code == 0
        assert output["data"]["format"] == "0.00%"

    def test_date_format(self, data_workbook: Path, tmp_path: Path):
        """Test date number format."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A5",
            "--number-format",
            "yyyy-mm-dd",
        )

        assert exit_code == 0
        assert output["data"]["format"] == "yyyy-mm-dd"

    def test_empty_format_error(self, data_workbook: Path, tmp_path: Path):
        """Test error on empty format string."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A5",
            "--number-format",
            " ",
        )

        assert exit_code == 1
