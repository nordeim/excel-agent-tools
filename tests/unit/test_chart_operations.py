"""Unit tests for chart operations (xls_add_chart).

Tests chart creation for all types (bar, line, pie, scatter),
data validation, positioning, and error handling.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(*args: str) -> tuple[dict, int]:
    """Helper to run xls_add_chart and parse output."""
    cmd = [sys.executable, "-m", "excel_agent.tools.objects.xls_add_chart", *args]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def chart_data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data suitable for charts."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "ChartData"

    # Headers
    ws["A1"] = "Category"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Q3"

    # Categories and data
    categories = ["Product A", "Product B", "Product C", "Product D"]
    for i, cat in enumerate(categories, start=2):
        ws[f"A{i}"] = cat
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = i * 200

    path = tmp_path / "chart_data.xlsx"
    wb.save(str(path))
    return path


class TestBarChart:
    """Test bar chart creation."""

    def test_create_bar_chart(self, chart_data_workbook: Path, tmp_path: Path):
        """Test creating a basic bar chart."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--categories-range",
            "A2:A5",
            "--position",
            "F2",
            "--title",
            "Sales by Quarter",
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["chart_type"] == "bar"
        assert output["data"]["title"] == "Sales by Quarter"

        # Verify chart exists
        wb = load_workbook(str(output_path))
        ws = wb.active
        assert len(ws._charts) == 1

    def test_bar_chart_without_categories(self, chart_data_workbook: Path, tmp_path: Path):
        """Test bar chart without explicit categories."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--position",
            "F2",
        )

        assert exit_code == 0


class TestLineChart:
    """Test line chart creation."""

    def test_create_line_chart(self, chart_data_workbook: Path, tmp_path: Path):
        """Test creating a line chart."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "line",
            "--data-range",
            "B1:D5",
            "--categories-range",
            "A2:A5",
            "--position",
            "F2",
            "--title",
            "Quarterly Trends",
        )

        assert exit_code == 0
        assert output["data"]["chart_type"] == "line"


class TestPieChart:
    """Test pie chart creation."""

    def test_create_pie_chart(self, chart_data_workbook: Path, tmp_path: Path):
        """Test creating a pie chart."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "pie",
            "--data-range",
            "A1:B5",
            "--position",
            "F2",
            "--title",
            "Sales Distribution",
        )

        assert exit_code == 0
        assert output["data"]["chart_type"] == "pie"


class TestScatterChart:
    """Test scatter chart creation."""

    def test_create_scatter_chart(self, chart_data_workbook: Path, tmp_path: Path):
        """Test creating a scatter chart."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "scatter",
            "--data-range",
            "B1:C5",
            "--position",
            "F2",
            "--title",
            "Correlation",
        )

        assert exit_code == 0
        assert output["data"]["chart_type"] == "scatter"


class TestChartValidation:
    """Test chart input validation."""

    def test_invalid_chart_type(self, chart_data_workbook: Path, tmp_path: Path):
        """Test that invalid chart types are rejected."""
        output_path = tmp_path / "output.xlsx"
        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "excel_agent.tools.objects.xls_add_chart",
                "--input",
                str(chart_data_workbook),
                "--output",
                str(output_path),
                "--type",
                "invalid",
                "--data-range",
                "B1:D5",
                "--position",
                "F2",
            ],
            capture_output=True,
            text=True,
        )

        # argparse should reject invalid choice
        assert result.returncode == 2

    def test_non_numeric_data_error(self, tmp_path: Path):
        """Test error on non-numeric data."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Text"
        ws["A2"] = "More Text"
        path = tmp_path / "text_data.xlsx"
        wb.save(str(path))

        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(path),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "A1:A2",
            "--position",
            "C1",
        )

        assert exit_code == 1
        assert "does not contain numeric data" in str(output.get("warnings", []))

    def test_invalid_position(self, chart_data_workbook: Path, tmp_path: Path):
        """Test error on invalid position."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--position",
            "Invalid",
        )

        assert exit_code == 1
        assert "Invalid position" in str(output.get("warnings", []))

    def test_invalid_data_range(self, chart_data_workbook: Path, tmp_path: Path):
        """Test error on invalid data range."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "InvalidRange",
            "--position",
            "F2",
        )

        assert exit_code == 1
        assert "Failed to parse" in str(output.get("warnings", []))

    def test_mismatched_categories(self, chart_data_workbook: Path, tmp_path: Path):
        """Test error when categories don't match data dimensions."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--categories-range",
            "A2:A3",  # Only 2 categories for 4 data rows
            "--position",
            "F2",
        )

        assert exit_code == 1
        assert "Counts must match" in str(output.get("warnings", []))


class TestChartStyling:
    """Test chart styling options."""

    def test_chart_with_style(self, chart_data_workbook: Path, tmp_path: Path):
        """Test chart with specific style."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--position",
            "F2",
            "--style",
            "10",
        )

        assert exit_code == 0
        assert output["data"]["style"] == 10

    def test_chart_dimensions(self, chart_data_workbook: Path, tmp_path: Path):
        """Test chart with custom dimensions."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(chart_data_workbook),
            "--output",
            str(output_path),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--position",
            "F2",
            "--width",
            "20",
            "--height",
            "15",
        )

        assert exit_code == 0
        assert output["data"]["width_cm"] == 20
        assert output["data"]["height_cm"] == 15
