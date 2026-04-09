"""Unit tests for table operations (xls_add_table).

Tests table creation, validation, style application, and error handling.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(*args: str) -> tuple[dict, int]:
    """Helper to run xls_add_table and parse output."""
    cmd = [sys.executable, "-m", "excel_agent.tools.objects.xls_add_table", *args]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data suitable for table conversion."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"

    # Data rows
    for i in range(2, 6):
        ws[f"A{i}"] = f"Product {i - 1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = f"=B{i}+C{i}"

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestTableCreation:
    """Test basic table creation."""

    def test_create_basic_table(self, data_workbook: Path, tmp_path: Path):
        """Test creating a basic table."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "SalesData",
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["table_name"] == "SalesData"
        assert output["data"]["row_count"] == 5
        assert output["data"]["column_count"] == 4

        # Verify table exists in file
        wb = load_workbook(str(output_path))
        ws = wb.active
        assert len(ws.tables) == 1
        table = list(ws.tables.values())[0]
        assert table.displayName == "SalesData"

    def test_table_with_style(self, data_workbook: Path, tmp_path: Path):
        """Test creating table with specific style."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "StyledTable",
            "--style",
            "TableStyleMedium2",
        )

        assert exit_code == 0
        wb = load_workbook(str(output_path))
        ws = wb.active
        table = list(ws.tables.values())[0]
        assert table.tableStyleInfo.name == "TableStyleMedium2"

    def test_table_with_different_sheets(self, data_workbook: Path, tmp_path: Path):
        """Test creating table on specific sheet."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--sheet",
            "Data",
            "--range",
            "A1:D5",
            "--name",
            "SheetSpecificTable",
        )

        assert exit_code == 0
        assert output["data"]["sheet"] == "Data"


class TestTableValidation:
    """Test table name and range validation."""

    def test_invalid_name_with_spaces(self, data_workbook: Path, tmp_path: Path):
        """Test that names with spaces are rejected."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "Invalid Name",
        )

        assert exit_code == 1
        assert output["status"] == "error"
        assert "Invalid table name" in str(output.get("warnings", []))

    def test_invalid_name_starts_with_number(self, data_workbook: Path, tmp_path: Path):
        """Test that names starting with numbers are rejected."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "123Table",
        )

        assert exit_code == 1
        assert output["status"] == "error"

    def test_invalid_style(self, data_workbook: Path, tmp_path: Path):
        """Test that invalid style names are rejected."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "TestTable",
            "--style",
            "InvalidStyle",
        )

        assert exit_code == 1
        assert "Invalid table style" in str(output.get("warnings", []))

    def test_duplicate_name(self, data_workbook: Path, tmp_path: Path):
        """Test that duplicate table names are rejected."""
        # First, create a table
        output_path = tmp_path / "output.xlsx"
        run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "DuplicateTable",
        )

        # Try to create another with same name
        output2, exit_code2 = run_tool(
            "--input",
            str(output_path),
            "--output",
            str(tmp_path / "output2.xlsx"),
            "--range",
            "A1:D5",
            "--name",
            "DuplicateTable",
        )

        assert exit_code2 == 1
        assert "already exists" in str(output2.get("warnings", []))

    def test_overlapping_ranges(self, data_workbook: Path, tmp_path: Path):
        """Test that overlapping table ranges are rejected."""
        # First table
        output_path = tmp_path / "output.xlsx"
        run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            "FirstTable",
        )

        # Try overlapping table
        output2, exit_code2 = run_tool(
            "--input",
            str(output_path),
            "--output",
            str(tmp_path / "output2.xlsx"),
            "--range",
            "A1:D3",
            "--name",
            "OverlapTable",
        )

        assert exit_code2 == 1
        warnings_str = str(output2.get("warnings", [])).lower()
        assert "overlaps" in warnings_str


class TestTableEdgeCases:
    """Test edge cases and error handling."""

    def test_single_row_warning(self, tmp_path: Path):
        """Test warning on single-row range."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Header"
        path = tmp_path / "single_row.xlsx"
        wb.save(str(path))

        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(path),
            "--output",
            str(output_path),
            "--range",
            "A1",
            "--name",
            "SingleRow",
        )

        # Should succeed but with warning
        assert exit_code == 0

    def test_missing_file(self, tmp_path: Path):
        """Test error on missing file."""
        output, exit_code = run_tool(
            "--input",
            str(tmp_path / "nonexistent.xlsx"),
            "--range",
            "A1:D5",
            "--name",
            "TestTable",
        )

        assert exit_code == 2

    def test_invalid_range_format(self, data_workbook: Path, tmp_path: Path):
        """Test error on invalid range format."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "InvalidRange",
            "--name",
            "TestTable",
        )

        assert exit_code == 1
        assert "Failed to parse range" in str(output.get("warnings", []))


class TestTableStyles:
    """Test different table styles."""

    @pytest.mark.parametrize(
        "style",
        [
            "TableStyleLight1",
            "TableStyleLight10",
            "TableStyleMedium5",
            "TableStyleMedium15",
            "TableStyleDark1",
            "TableStyleDark5",
        ],
    )
    def test_various_styles(self, data_workbook: Path, tmp_path: Path, style: str):
        """Test various table styles."""
        output_path = tmp_path / f"output_{style}.xlsx"
        output, exit_code = run_tool(
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:D5",
            "--name",
            f"Table{style}",
            "--style",
            style,
        )

        assert exit_code == 0, f"Style {style} should be valid"
        wb = load_workbook(str(output_path))
        ws = wb.active
        table = list(ws.tables.values())[0]
        assert table.tableStyleInfo.name == style
