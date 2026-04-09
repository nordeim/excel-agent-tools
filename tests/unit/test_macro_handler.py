"""Unit tests for the macro_handler module.

Tests the OleToolsMacroAnalyzer and related functionality.
Uses mocks for oletools where appropriate to avoid requiring actual macro files.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from excel_agent.core.macro_handler import (
    MacroAnalysisResult,
    MacroModule,
    OleToolsMacroAnalyzer,
    SUSPICIOUS_PATTERNS,
    get_analyzer,
    has_macros,
)

# -----------------------------------------------------------------------------
# MacroModule Tests
# -----------------------------------------------------------------------------


def test_macro_module_creation():
    """Test creating a MacroModule instance."""
    module = MacroModule(
        name="Module1",
        code='Sub Test()\nMsgBox "Hello"\nEnd Sub',
        is_stream=False,
        risk_indicators=["auto_exec"],
    )
    assert module.name == "Module1"
    assert "MsgBox" in module.code
    assert module.risk_indicators == ["auto_exec"]


def test_macro_module_defaults():
    """Test MacroModule with default values."""
    module = MacroModule(name="Empty", code="")
    assert module.is_stream is False
    assert module.risk_indicators == []


# -----------------------------------------------------------------------------
# MacroAnalysisResult Tests
# -----------------------------------------------------------------------------


def test_macro_analysis_result_defaults():
    """Test MacroAnalysisResult default values."""
    result = MacroAnalysisResult()
    assert result.has_macros is False
    assert result.is_signed is False
    assert result.signature_valid is False
    assert result.module_count == 0
    assert result.modules == []
    assert result.risk_score == 0
    assert result.risk_level == "none"
    assert result.auto_exec_functions == []
    assert result.iocs == []
    assert result.errors == []


def test_macro_analysis_result_to_dict():
    """Test converting MacroAnalysisResult to dict."""
    result = MacroAnalysisResult(
        has_macros=True,
        module_count=2,
        risk_score=50,
        risk_level="medium",
        auto_exec_functions=["AutoOpen"],
    )
    result.modules.append(MacroModule(name="Test", code="Sub AutoOpen()\nEnd Sub"))

    d = result.to_dict()
    assert d["has_macros"] is True
    assert d["module_count"] == 2
    assert d["risk_score"] == 50
    assert d["risk_level"] == "medium"
    assert d["auto_exec_functions"] == ["AutoOpen"]
    assert len(d["modules"]) == 1


def test_macro_analysis_result_to_dict_with_code_preview():
    """Test that code_preview is truncated to 500 chars."""
    long_code = "Sub Test()\n" + 'MsgBox "Hello"\n' * 100 + "End Sub"
    result = MacroAnalysisResult(has_macros=True)
    result.modules.append(MacroModule(name="LongCode", code=long_code))

    d = result.to_dict()
    preview = d["modules"][0]["code_preview"]
    assert len(preview) <= 500
    assert preview.startswith("Sub Test()")


# -----------------------------------------------------------------------------
# Suspicious Patterns Tests
# -----------------------------------------------------------------------------


def test_suspicious_patterns_defined():
    """Test that suspicious patterns are defined."""
    assert "auto_exec" in SUSPICIOUS_PATTERNS
    assert "shell" in SUSPICIOUS_PATTERNS
    assert "network" in SUSPICIOUS_PATTERNS
    assert "obfuscation" in SUSPICIOUS_PATTERNS

    # Check auto_exec patterns
    auto_exec = SUSPICIOUS_PATTERNS["auto_exec"]
    assert "AutoOpen" in auto_exec
    assert "Workbook_Open" in auto_exec

    # Check shell patterns
    shell = SUSPICIOUS_PATTERNS["shell"]
    assert "Shell" in shell
    assert "CreateObject" in shell


# -----------------------------------------------------------------------------
# has_macros Quick Check Tests
# -----------------------------------------------------------------------------


def test_has_macros_returns_false_for_nonexistent_file(tmp_path: Path):
    """Test has_macros returns False for non-existent file."""
    path = tmp_path / "nonexistent.xlsx"
    assert has_macros(path) is False


def test_has_macros_returns_false_for_xlsx(tmp_path: Path):
    """Test has_macros returns False for clean .xlsx file."""
    wb = Workbook()
    path = tmp_path / "clean.xlsx"
    wb.save(str(path))
    assert has_macros(path) is False


def test_has_macros_returns_true_for_xlsm_with_vba(tmp_path: Path):
    """Test has_macros returns True for .xlsm with vbaProject.bin marker."""
    wb = Workbook()
    path = tmp_path / "with_macros.xlsm"
    wb.save(str(path))

    # Inject vbaProject.bin marker
    import zipfile

    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA_PROJECT_DATA")

    assert has_macros(path) is True


def test_has_macros_handles_bad_zip(tmp_path: Path):
    """Test has_macros handles corrupt zip files gracefully."""
    path = tmp_path / "corrupt.zip"
    path.write_bytes(b"Not a zip file")
    assert has_macros(path) is False


# -----------------------------------------------------------------------------
# OleToolsMacroAnalyzer Tests (without actual oletools)
# -----------------------------------------------------------------------------


def test_analyzer_creation():
    """Test creating OleToolsMacroAnalyzer."""
    analyzer = OleToolsMacroAnalyzer()
    assert analyzer is not None


def test_analyzer_returns_no_macros_for_clean_file(tmp_path: Path):
    """Test analyzer returns no macros for clean .xlsx."""
    wb = Workbook()
    path = tmp_path / "clean.xlsx"
    wb.save(str(path))

    analyzer = OleToolsMacroAnalyzer()
    result = analyzer.analyze(path)

    assert result.has_macros is False
    assert result.risk_level == "none"


def test_analyzer_returns_error_for_nonexistent_file(tmp_path: Path):
    """Test analyzer returns error for non-existent file."""
    path = tmp_path / "nonexistent.xlsx"
    analyzer = OleToolsMacroAnalyzer()
    result = analyzer.analyze(path)

    assert result.has_macros is False
    assert len(result.errors) == 1
    assert "File not found" in result.errors[0]


def test_analyzer_detects_vba_presence(tmp_path: Path):
    """Test analyzer detects VBA presence via zip inspection."""
    wb = Workbook()
    path = tmp_path / "with_macros.xlsm"
    wb.save(str(path))

    # Inject vbaProject.bin
    import zipfile

    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA_DATA")

    analyzer = OleToolsMacroAnalyzer()
    result = analyzer.analyze(path)

    assert result.has_macros is True


@patch("excel_agent.core.macro_handler.OleToolsMacroAnalyzer._check_vba_presence")
@patch("excel_agent.core.macro_handler.OleToolsMacroAnalyzer._analyze_risk")
@patch("excel_agent.core.macro_handler.OleToolsMacroAnalyzer._calculate_risk_score")
@patch("excel_agent.core.macro_handler.OleToolsMacroAnalyzer._get_risk_level")
@patch("excel_agent.core.macro_handler.OleToolsMacroAnalyzer._find_auto_exec")
def test_analyzer_full_workflow(
    mock_find_auto_exec,
    mock_get_risk_level,
    mock_calc_score,
    mock_analyze_risk,
    mock_check_vba,
    tmp_path: Path,
):
    """Test analyzer full workflow with mocked oletools."""
    mock_check_vba.return_value = True
    mock_calc_score.return_value = 75
    mock_get_risk_level.return_value = "high"
    mock_find_auto_exec.return_value = ["Module1: AutoOpen"]

    wb = Workbook()
    path = tmp_path / "test.xlsm"
    wb.save(str(path))

    # Create a mock module
    mock_module = MagicMock()
    mock_module.code = "Sub AutoOpen()\nEnd Sub"

    analyzer = OleToolsMacroAnalyzer()

    # Mock the oletools import to None (simulate unavailable)
    analyzer._olevba = None

    result = analyzer.analyze(path)

    # Should return basic result without detailed analysis
    assert result.has_macros is True
    assert "oletools not available" in result.errors[0]


def test_analyze_risk_detects_auto_exec():
    """Test _analyze_risk detects auto-exec patterns."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(name="Test", code='Sub AutoOpen()\nMsgBox "Hi"\nEnd Sub')

    analyzer._analyze_risk(module)

    assert "auto_exec" in module.risk_indicators


def test_analyze_risk_detects_shell():
    """Test _analyze_risk detects shell patterns."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(name="Test", code='Shell "cmd.exe"\nCreateObject "WScript.Shell"')

    analyzer._analyze_risk(module)

    assert "shell" in module.risk_indicators


def test_analyze_risk_detects_network():
    """Test _analyze_risk detects network patterns."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(
        name="Test", code='WinHttp.Open URLDownloadToFile CreateObject("MSXML2.XMLHTTP")'
    )

    analyzer._analyze_risk(module)

    assert "network" in module.risk_indicators


def test_analyze_risk_detects_obfuscation():
    """Test _analyze_risk detects obfuscation patterns."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(name="Test", code='Chr(65) & Chr(66) &H41 StrReverse("cba")')

    analyzer._analyze_risk(module)

    assert "obfuscation" in module.risk_indicators


def test_analyze_risk_removes_duplicates():
    """Test _analyze_risk removes duplicate indicators."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(
        name="Test",
        code="AutoOpen\nAutoOpen",  # Duplicate patterns
        risk_indicators=[],  # type: ignore
    )

    analyzer._analyze_risk(module)

    # Should only have auto_exec once
    assert module.risk_indicators.count("auto_exec") == 1


# -----------------------------------------------------------------------------
# Risk Score Calculation Tests
# -----------------------------------------------------------------------------


def test_calculate_risk_score_zero():
    """Test risk score is 0 for no macros."""
    analyzer = OleToolsMacroAnalyzer()
    result = MacroAnalysisResult(has_macros=False, modules=[])
    score = analyzer._calculate_risk_score(result)
    assert score == 0


def test_calculate_risk_score_base():
    """Test base score for having macros."""
    analyzer = OleToolsMacroAnalyzer()
    result = MacroAnalysisResult(has_macros=True, modules=[MacroModule(name="M1", code="")])
    score = analyzer._calculate_risk_score(result)
    assert score == 10  # Base score for having macros


def test_calculate_risk_score_with_indicators():
    """Test risk score with risk indicators."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(name="M1", code="", risk_indicators=["auto_exec", "shell"])
    result = MacroAnalysisResult(has_macros=True, modules=[module])
    score = analyzer._calculate_risk_score(result)

    # Base 10 + (2 indicators * 15) + 25 (auto_exec) + 20 (shell)
    expected = 10 + 30 + 25 + 20
    assert score == expected


def test_calculate_risk_score_capped_at_100():
    """Test risk score is capped at 100."""
    analyzer = OleToolsMacroAnalyzer()
    module = MacroModule(
        name="M1",
        code="",
        risk_indicators=["auto_exec", "shell", "network", "obfuscation"] * 10,
    )
    result = MacroAnalysisResult(has_macros=True, modules=[module])
    score = analyzer._calculate_risk_score(result)

    assert score == 100  # Capped


def test_get_risk_level_none():
    """Test risk level for score 0."""
    analyzer = OleToolsMacroAnalyzer()
    assert analyzer._get_risk_level(0) == "none"


def test_get_risk_level_low():
    """Test risk level for low score."""
    analyzer = OleToolsMacroAnalyzer()
    assert analyzer._get_risk_level(10) == "low"
    assert analyzer._get_risk_level(24) == "low"


def test_get_risk_level_medium():
    """Test risk level for medium score."""
    analyzer = OleToolsMacroAnalyzer()
    assert analyzer._get_risk_level(25) == "medium"
    assert analyzer._get_risk_level(49) == "medium"


def test_get_risk_level_high():
    """Test risk level for high score."""
    analyzer = OleToolsMacroAnalyzer()
    assert analyzer._get_risk_level(50) == "high"
    assert analyzer._get_risk_level(100) == "high"


# -----------------------------------------------------------------------------
# Auto-Exec Detection Tests
# -----------------------------------------------------------------------------


def test_find_auto_exec_detects_patterns():
    """Test _find_auto_exec finds auto-exec functions."""
    analyzer = OleToolsMacroAnalyzer()
    result = MacroAnalysisResult(
        has_macros=True,
        modules=[
            MacroModule(name="Module1", code="Sub AutoOpen()\nEnd Sub"),
            MacroModule(name="Module2", code="Sub Workbook_Open()\nEnd Sub"),
        ],
    )

    auto_exec = analyzer._find_auto_exec(result)

    assert len(auto_exec) == 2
    assert "Module1: AutoOpen" in auto_exec
    assert "Module2: Workbook_Open" in auto_exec


def test_find_auto_exec_empty():
    """Test _find_auto_exec returns empty list for no auto-exec."""
    analyzer = OleToolsMacroAnalyzer()
    result = MacroAnalysisResult(
        has_macros=True,
        modules=[MacroModule(name="Module1", code="Sub Normal()\nEnd Sub")],
    )

    auto_exec = analyzer._find_auto_exec(result)

    assert auto_exec == []


# -----------------------------------------------------------------------------
# get_analyzer Factory Tests
# -----------------------------------------------------------------------------


def test_get_analyzer_returns_ole_tools_analyzer():
    """Test get_analyzer returns OleToolsMacroAnalyzer."""
    analyzer = get_analyzer()
    assert isinstance(analyzer, OleToolsMacroAnalyzer)


# -----------------------------------------------------------------------------
# Error Handling Tests
# -----------------------------------------------------------------------------


def test_analyzer_handles_exception_gracefully(tmp_path: Path):
    """Test analyzer handles exceptions gracefully."""
    wb = Workbook()
    path = tmp_path / "test.xlsm"
    wb.save(str(path))

    # Inject vba to trigger presence check
    import zipfile

    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA")

    analyzer = OleToolsMacroAnalyzer()

    # Mock oletools to raise exception
    mock_vba = MagicMock()
    mock_vba.VBA_Parser.side_effect = Exception("Mock oletools error")
    analyzer._olevba = mock_vba

    result = analyzer.analyze(path)

    assert result.has_macros is True  # Detected via zip
    assert len(result.errors) == 1
    assert "Analysis error" in result.errors[0]
