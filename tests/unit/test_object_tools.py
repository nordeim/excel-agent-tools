"""Unit tests for object tools (images, comments, data validation).

Tests xls_add_image, xls_add_comment, and xls_set_data_validation.
"""

from __future__ import annotations

import subprocess
import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run a tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.objects.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"raw_output": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def simple_workbook(tmp_path: Path) -> Path:
    """Create a simple workbook for testing."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Test"
    ws["B1"] = 100
    path = tmp_path / "simple.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def test_image(tmp_path: Path) -> Path:
    """Create a test PNG image."""
    img_path = tmp_path / "test.png"
    img = PILImage.new("RGB", (200, 100), color="red")
    img.save(str(img_path))
    return img_path


class TestAddImage:
    """Test xls_add_image tool."""

    def test_add_png_image(self, simple_workbook: Path, test_image: Path, tmp_path: Path):
        """Test adding a PNG image."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(test_image),
            "--position",
            "D1",
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["position"] == "D1"
        assert output["data"]["original_width"] == 200
        assert output["data"]["original_height"] == 100

    def test_image_resize_width(self, simple_workbook: Path, test_image: Path, tmp_path: Path):
        """Test resizing image by width only."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(test_image),
            "--position",
            "D1",
            "--width",
            "100",
        )

        assert exit_code == 0
        # Should maintain aspect ratio
        assert output["data"]["final_width"] == 100
        assert output["data"]["final_height"] == 50  # Half height due to aspect ratio

    def test_image_resize_height(self, simple_workbook: Path, test_image: Path, tmp_path: Path):
        """Test resizing image by height only."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(test_image),
            "--position",
            "D1",
            "--height",
            "50",
        )

        assert exit_code == 0
        # Should maintain aspect ratio
        assert output["data"]["final_height"] == 50
        assert output["data"]["final_width"] == 100  # Half width due to aspect ratio

    def test_image_not_found(self, simple_workbook: Path, tmp_path: Path):
        """Test error when image file doesn't exist."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(tmp_path / "nonexistent.png"),
            "--position",
            "D1",
        )

        assert exit_code == 2
        assert "not found" in str(output.get("warnings", [])).lower()

    def test_unsupported_format(self, simple_workbook: Path, tmp_path: Path):
        """Test error on unsupported image format."""
        # Create a fake file with unsupported extension
        fake_img = tmp_path / "test.tiff"
        fake_img.write_bytes(b"fake image data")

        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(fake_img),
            "--position",
            "D1",
        )

        assert exit_code == 1
        assert "Unsupported image format" in str(output.get("warnings", []))


class TestAddComment:
    """Test xls_add_comment tool."""

    def test_add_basic_comment(self, simple_workbook: Path, tmp_path: Path):
        """Test adding a basic comment."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--cell",
            "B1",
            "--text",
            "This is a comment",
            "--author",
            "TestUser",
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["cell"] == "B1"
        assert output["data"]["author"] == "TestUser"

        # Verify comment exists
        wb = load_workbook(str(output_path))
        ws = wb.active
        assert ws["B1"].comment is not None
        assert ws["B1"].comment.text == "This is a comment"
        assert ws["B1"].comment.author == "TestUser"

    def test_add_comment_default_author(self, simple_workbook: Path, tmp_path: Path):
        """Test adding comment with default author."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--cell",
            "A1",
            "--text",
            "Test comment",
        )

        assert exit_code == 0
        assert output["data"]["author"] == "excel-agent"

    def test_comment_replacement_warning(self, simple_workbook: Path, tmp_path: Path):
        """Test warning when replacing existing comment."""
        # First add a comment
        wb = load_workbook(str(simple_workbook))
        from openpyxl.comments import Comment

        ws = wb.active
        ws["A1"].comment = Comment("Original", "Author")
        wb.save(str(simple_workbook))

        # Now replace it
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--cell",
            "A1",
            "--text",
            "Replacement comment",
        )

        assert exit_code == 0
        assert "already has a comment" in str(output.get("warnings", []))

    def test_empty_comment_error(self, simple_workbook: Path, tmp_path: Path):
        """Test error on empty comment text."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--cell",
            "A1",
            "--text",
            "   ",  # Whitespace only
        )

        assert exit_code == 1
        assert "cannot be empty" in str(output.get("warnings", []))

    def test_invalid_cell(self, simple_workbook: Path, tmp_path: Path):
        """Test error on invalid cell reference."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--cell",
            "Invalid",
            "--text",
            "Test",
        )

        assert exit_code == 1
        assert "Invalid cell" in str(output.get("warnings", []))


class TestDataValidation:
    """Test xls_set_data_validation tool."""

    def test_list_validation(self, simple_workbook: Path, tmp_path: Path):
        """Test list data validation."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A10",
            "--type",
            "list",
            "--formula1",
            '"Option1,Option2,Option3"',
        )

        assert exit_code == 0
        assert output["status"] == "success"
        assert output["data"]["validation_type"] == "list"

    def test_whole_number_validation(self, simple_workbook: Path, tmp_path: Path):
        """Test whole number validation."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "B1:B10",
            "--type",
            "whole",
            "--formula1",
            "0",
            "--operator",
            "greaterThanOrEqual",
        )

        assert exit_code == 0
        assert output["data"]["validation_type"] == "whole"

    def test_decimal_validation(self, simple_workbook: Path, tmp_path: Path):
        """Test decimal number validation."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "B1:B10",
            "--type",
            "decimal",
            "--formula1",
            "0.0",
            "--formula2",
            "100.0",
            "--operator",
            "between",
        )

        assert exit_code == 0
        assert output["data"]["validation_type"] == "decimal"

    def test_text_length_validation(self, simple_workbook: Path, tmp_path: Path):
        """Test text length validation."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A10",
            "--type",
            "textLength",
            "--formula1",
            "1",
            "--formula2",
            "50",
            "--operator",
            "between",
        )

        assert exit_code == 0
        assert output["data"]["validation_type"] == "textLength"

    def test_custom_validation(self, simple_workbook: Path, tmp_path: Path):
        """Test custom formula validation."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A10",
            "--type",
            "custom",
            "--formula1",
            "=ISNUMBER(A1)",
        )

        assert exit_code == 0
        assert output["data"]["validation_type"] == "custom"

    def test_invalid_validation_type(self, simple_workbook: Path, tmp_path: Path):
        """Test error on invalid validation type."""
        output_path = tmp_path / "output.xlsx"
        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "excel_agent.tools.objects.xls_set_data_validation",
                "--input",
                str(simple_workbook),
                "--output",
                str(output_path),
                "--range",
                "A1:A10",
                "--type",
                "invalid",
                "--formula1",
                "test",
            ],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 2  # argparse error

    def test_error_message(self, simple_workbook: Path, tmp_path: Path):
        """Test custom error message."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A10",
            "--type",
            "whole",
            "--formula1",
            "0",
            "--show-error",
            "--error-title",
            "Invalid Value",
            "--error-message",
            "Please enter a positive number",
        )

        assert exit_code == 0

    def test_input_message(self, simple_workbook: Path, tmp_path: Path):
        """Test input message display."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(simple_workbook),
            "--output",
            str(output_path),
            "--range",
            "A1:A10",
            "--type",
            "list",
            "--formula1",
            '"Yes,No"',
            "--show-input",
            "--input-title",
            "Selection",
            "--input-message",
            "Choose Yes or No",
        )

        assert exit_code == 0
