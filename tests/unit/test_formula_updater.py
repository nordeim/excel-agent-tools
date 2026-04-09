"""Tests for the formula reference updating engine."""

from __future__ import annotations

from openpyxl import Workbook

from excel_agent.core.formula_updater import (
    adjust_col_references,
    adjust_row_references,
    rename_sheet_in_formulas,
)


class TestRenameSheetInFormulas:
    def test_simple_cross_sheet_ref(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Data"
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "=Data!B1"

        updated = rename_sheet_in_formulas(wb, "Data", "RawData")
        assert updated >= 1
        assert ws2["A1"].value == "=RawData!B1"

    def test_quoted_sheet_name(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "My Data"
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "='My Data'!B1"

        updated = rename_sheet_in_formulas(wb, "My Data", "New Data")
        assert updated >= 1
        assert "'New Data'!B1" in ws2["A1"].value

    def test_no_refs_to_rename(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = "=B1+1"

        updated = rename_sheet_in_formulas(wb, "Sheet1", "NewName")
        # Local refs (no sheet prefix) should NOT be changed
        assert updated == 0
        assert ws["A1"].value == "=B1+1"

    def test_multiple_refs_in_one_formula(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Src"
        ws2 = wb.create_sheet("Calc")
        ws2["A1"] = "=Src!A1+Src!B1"

        updated = rename_sheet_in_formulas(wb, "Src", "Source")
        assert updated >= 1
        assert ws2["A1"].value == "=Source!A1+Source!B1"


class TestAdjustRowReferences:
    def test_insert_shifts_down(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws["A1"] = 10
        ws["A5"] = "=A1+1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!A5"

        updated = adjust_row_references(wb, "Sheet1", 3, 2)
        assert updated >= 1
        # A5 → A7 (shifted by 2 because insert at row 3)
        assert "A7" in ws2["A1"].value

    def test_delete_shifts_up(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!A10"

        updated = adjust_row_references(wb, "Sheet1", 5, -2)
        assert updated >= 1
        assert "A8" in ws2["A1"].value

    def test_ref_above_start_not_shifted(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!A2"

        updated = adjust_row_references(wb, "Sheet1", 5, 3)
        # A2 is above start_row=5, should not shift
        assert "A2" in ws2["A1"].value


class TestAdjustColReferences:
    def test_insert_shifts_right(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!D1"

        updated = adjust_col_references(wb, "Sheet1", 2, 1)
        assert updated >= 1
        # D (col 4) → E (col 5) because insert at col 2
        assert "E1" in ws2["A1"].value

    def test_ref_before_start_not_shifted(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "=Sheet1!A1"

        updated = adjust_col_references(wb, "Sheet1", 3, 2)
        # A1 is before start_col=3, should not shift
        assert "A1" in ws2["A1"].value
