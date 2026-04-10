"""Benchmark large dataset read performance."""

from __future__ import annotations

import tempfile
import time
from pathlib import Path

import pytest
from openpyxl import Workbook


def create_large_workbook(rows: int, cols: int) -> Path:
    """Create a large workbook for benchmarking."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")

    # Header
    ws.append([f"Col{i}" for i in range(1, cols + 1)])

    # Data rows
    for i in range(1, rows + 1):
        ws.append([i, f"Item {i}", i * 1.5, i * 2, i % 2 == 0])

    path = Path(tempfile.mktemp(suffix=".xlsx"))
    wb.save(str(path))
    return path


def bench_read_large() -> dict[str, float]:
    """Benchmark reading 100k rows."""
    rows, cols = 100_000, 5
    path = create_large_workbook(rows, cols)

    try:
        start = time.perf_counter()

        # Simulate reading with chunked I/O
        from excel_agent.core.chunked_io import read_range_chunked
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active

        total_cells = 0
        for chunk in read_range_chunked(ws, 1, 1, rows, cols, chunk_size=10_000):
            for row in chunk:
                total_cells += len(row)

        elapsed = time.perf_counter() - start

        return {
            "rows": rows,
            "cols": cols,
            "elapsed_seconds": round(elapsed, 2),
            "cells_per_second": round(total_cells / elapsed, 0),
            "target_met": elapsed < 3.0,
        }
    finally:
        path.unlink(missing_ok=True)


if __name__ == "__main__":
    results = bench_read_large()
    print(f"Read {results['rows']} rows in {results['elapsed_seconds']}s")
    print(f"Cells/sec: {results['cells_per_second']}")
    print(f"Target met (<3s): {results['target_met']}")
