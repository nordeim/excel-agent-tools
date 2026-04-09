"""Integration tests for structural mutation tools via subprocess."""

from __future__ import annotations

import json
import subprocess  # noqa: S603
import sys
from pathlib import Path

from openpyxl import load_workbook


def _run_tool(tool_module: str, *args: str) -> tuple[dict, int]:
    """Run a CLI tool and return (parsed_json, return_code)."""
    result = subprocess.run(  # noqa: S603
        [sys.executable, "-m", f"excel_agent.tools.{tool_module}", *args],
        capture_output=True,
        text=True,
        timeout=30,
    )
    data = json.loads(result.stdout) if result.stdout.strip() else {}
    return data, result.returncode


class TestAddSheet:
    def test_add_at_end(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_add_sheet",
            "--input",
            str(work),
            "--output",
            str(work),
            "--name",
            "NewSheet",
        )
        assert code == 0
        wb = load_workbook(str(work))
        assert "NewSheet" in wb.sheetnames

    def test_add_at_position(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_add_sheet",
            "--input",
            str(work),
            "--output",
            str(work),
            "--name",
            "First",
            "--position",
            "0",
        )
        assert code == 0
        wb = load_workbook(str(work))
        assert wb.sheetnames[0] == "First"


class TestDeleteSheet:
    def test_without_token_fails(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work),
            "--output",
            str(work),
            "--name",
            "Sheet3",
        )
        assert code == 1  # ValidationError for missing token


class TestMoveSheet:
    def test_move_to_front(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_move_sheet",
            "--input",
            str(work),
            "--output",
            str(work),
            "--name",
            "Sheet3",
            "--position",
            "0",
        )
        assert code == 0
        wb = load_workbook(str(work))
        assert wb.sheetnames[0] == "Sheet3"


class TestInsertRows:
    def test_insert_rows(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        wb_before = load_workbook(str(work))
        max_row_before = wb_before["Sheet1"].max_row

        data, code = _run_tool(
            "structure.xls_insert_rows",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--before-row",
            "3",
            "--count",
            "2",
        )
        assert code == 0

        wb_after = load_workbook(str(work))
        assert wb_after["Sheet1"].max_row >= max_row_before + 2


class TestDeleteRows:
    def test_without_token_fails(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_delete_rows",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--start-row",
            "5",
            "--count",
            "1",
        )
        assert code == 1  # Missing token


class TestInsertColumns:
    def test_insert_columns(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        data, code = _run_tool(
            "structure.xls_insert_columns",
            "--input",
            str(work),
            "--output",
            str(work),
            "--sheet",
            "Sheet1",
            "--before-column",
            "B",
            "--count",
            "1",
        )
        assert code == 0
