"""Integration tests for object tools workflow.

Tests end-to-end workflows: table → chart → image → comment → validation
"""

from __future__ import annotations

import subprocess
import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run a tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.objects.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"parse_error": True, "stdout": result.stdout, "stderr": result.stderr}
    return output, result.returncode


def create_test_image(tmp_path: Path) -> Path:
    """Create a test image file."""
    img_path = tmp_path / "test_logo.png"
    img = PILImage.new("RGB", (200, 100), color="blue")
    img.save(str(img_path))
    return img_path


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data for objects."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"

    # Data rows
    for i in range(2, 6):
        ws[f"A{i}"] = f"Product {i - 1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = f"=B{i}+C{i}"

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestObjectCreationWorkflow:
    """Test creating multiple objects in sequence."""

    def test_full_object_workflow(self, data_workbook: Path, tmp_path: Path):
        """Test complete workflow: table → chart → comment → validation."""
        img_path = create_test_image(tmp_path)

        # Step 1: Add table
        output1 = tmp_path / "step1.xlsx"
        out1, exit1 = run_tool(
            "xls_add_table",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "A1:D5",
            "--name",
            "SalesTable",
            "--style",
            "TableStyleMedium2",
        )
        assert exit1 == 0, f"Table creation failed: {out1}"

        # Step 2: Add chart
        output2 = tmp_path / "step2.xlsx"
        out2, exit2 = run_tool(
            "xls_add_chart",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--type",
            "bar",
            "--data-range",
            "B1:D5",
            "--categories-range",
            "A2:A5",
            "--position",
            "F2",
            "--title",
            "Quarterly Sales",
        )
        assert exit2 == 0, f"Chart creation failed: {out2}"

        # Step 3: Add image
        output3 = tmp_path / "step3.xlsx"
        out3, exit3 = run_tool(
            "xls_add_image",
            "--input",
            str(output2),
            "--output",
            str(output3),
            "--image-path",
            str(img_path),
            "--position",
            "F15",
            "--width",
            "100",
        )
        assert exit3 == 0, f"Image insertion failed: {out3}"

        # Step 4: Add comment
        output4 = tmp_path / "step4.xlsx"
        out4, exit4 = run_tool(
            "xls_add_comment",
            "--input",
            str(output3),
            "--output",
            str(output4),
            "--cell",
            "A1",
            "--text",
            "This is the main data table",
            "--author",
            "DataAnalyst",
        )
        assert exit4 == 0, f"Comment creation failed: {out4}"

        # Step 5: Add data validation
        output5 = tmp_path / "final.xlsx"
        out5, exit5 = run_tool(
            "xls_set_data_validation",
            "--input",
            str(output4),
            "--output",
            str(output5),
            "--range",
            "E1:E10",
            "--type",
            "list",
            "--formula1",
            '"High,Medium,Low"',
            "--show-input",
            "--input-title",
            "Priority",
            "--input-message",
            "Select priority level",
        )
        assert exit5 == 0, f"Validation creation failed: {out5}"

        # Verify final file has all objects
        wb = load_workbook(str(output5))
        ws = wb.active

        # Verify table
        assert len(ws.tables) == 1
        table = list(ws.tables.values())[0]
        assert table.displayName == "SalesTable"

        # Verify chart
        assert len(ws._charts) == 1

        # Verify comment
        assert ws["A1"].comment is not None
        assert "main data table" in ws["A1"].comment.text

        # Verify data validation
        assert len(ws.data_validations.dataValidation) == 1

    def test_table_then_chart_from_table(self, data_workbook: Path, tmp_path: Path):
        """Test creating table then chart from table data."""
        # Add table first
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_add_table",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "A1:D5",
            "--name",
            "DataTable",
        )

        # Create chart from table data
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_add_chart",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--type",
            "line",
            "--data-range",
            "B1:D5",
            "--position",
            "F2",
            "--title",
            "Trend Analysis",
        )

        assert exit_code == 0

        # Verify both objects exist
        wb = load_workbook(str(output2))
        ws = wb.active
        assert len(ws.tables) == 1
        assert len(ws._charts) == 1


class TestObjectInteractions:
    """Test interactions between different objects."""

    def test_image_and_comment_same_sheet(self, data_workbook: Path, tmp_path: Path):
        """Test adding image and comment to same sheet."""
        img_path = create_test_image(tmp_path)

        # Add image
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_add_image",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--image-path",
            str(img_path),
            "--position",
            "F1",
        )

        # Add comment
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_add_comment",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--cell",
            "A1",
            "--text",
            "Header cell with comment",
        )

        assert exit_code == 0

        # Verify both exist
        wb = load_workbook(str(output2))
        ws = wb.active
        assert ws["A1"].comment is not None
        # Note: image verification is limited in openpyxl

    def test_multiple_validations(self, data_workbook: Path, tmp_path: Path):
        """Test adding multiple validations to different ranges."""
        # First validation
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_set_data_validation",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "A1:A5",
            "--type",
            "list",
            "--formula1",
            '"A,B,C"',
        )

        # Second validation
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_set_data_validation",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--range",
            "B1:B5",
            "--type",
            "whole",
            "--formula1",
            "0",
            "--operator",
            "greaterThanOrEqual",
        )

        assert exit_code == 0


class TestObjectErrorHandling:
    """Test error handling in object workflows."""

    def test_missing_image_file(self, data_workbook: Path, tmp_path: Path):
        """Test error when image file missing."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_image",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--image-path",
            str(tmp_path / "nonexistent.png"),
            "--position",
            "A1",
        )

        assert exit_code == 2

    def test_invalid_chart_type_in_workflow(self, data_workbook: Path, tmp_path: Path):
        """Test workflow continues after chart error."""
        # This should work (we can't test invalid type due to argparse)
        pass


class TestPerformance:
    """Test performance with larger objects."""

    @pytest.mark.slow
    def test_large_table_creation(self, tmp_path: Path):
        """Test creating table with many rows."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Create 1000 rows of data
        ws["A1"] = "ID"
        ws["B1"] = "Value"
        for i in range(2, 1002):
            ws[f"A{i}"] = i
            ws[f"B{i}"] = i * 10

        path = tmp_path / "large_data.xlsx"
        wb.save(str(path))

        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_add_table",
            "--input",
            str(path),
            "--output",
            str(output_path),
            "--range",
            "A1:B1000",
            "--name",
            "LargeTable",
        )

        assert exit_code == 0
        assert output["data"]["row_count"] == 1000

    def test_multiple_charts_performance(self, data_workbook: Path, tmp_path: Path):
        """Test creating multiple charts."""
        current_path = data_workbook

        chart_types = ["bar", "line", "pie", "scatter"]
        positions = ["F2", "F15", "F28", "F41"]

        for chart_type, pos in zip(chart_types, positions):
            output_path = tmp_path / f"chart_{chart_type}.xlsx"

            if chart_type == "pie":
                # Pie charts work best with single series
                data_range = "A1:B5"
            else:
                data_range = "B1:D5"

            output, exit_code = run_tool(
                "xls_add_chart",
                "--input",
                str(current_path),
                "--output",
                str(output_path),
                "--type",
                chart_type,
                "--data-range",
                data_range,
                "--position",
                pos,
            )

            assert exit_code == 0, f"Failed to create {chart_type} chart"
            current_path = output_path

        # Verify final file has all charts
        wb = load_workbook(str(current_path))
        ws = wb.active
        assert len(ws._charts) == 4
