"""Integration tests for formatting workflow.

Tests end-to-end workflows combining multiple formatting operations.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def run_tool(tool: str, *args: str) -> tuple[dict, int]:
    """Helper to run a formatting tool and parse output."""
    cmd = [
        sys.executable,
        "-m",
        f"excel_agent.tools.formatting.{tool}",
        *args,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    try:
        import json

        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"parse_error": True, "stdout": result.stdout, "stderr": result.stderr}
    return output, result.returncode


@pytest.fixture
def data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with data for formatting tests."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Q1 Sales"
    ws["C1"] = "Q2 Sales"
    ws["D1"] = "Total"

    # Data rows
    for i in range(2, 11):
        ws[f"A{i}"] = f"Product {i - 1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = i * 150
        ws[f"D{i}"] = i * 250

    path = tmp_path / "data.xlsx"
    wb.save(str(path))
    return path


class TestFullFormattingWorkflow:
    """Test complete formatting workflow."""

    def test_full_formatting_workflow(self, data_workbook: Path, tmp_path: Path):
        """Test complete workflow: format range → column width → freeze → conditional format → number format."""
        # Step 1: Format headers
        output1 = tmp_path / "step1.xlsx"
        out1, exit1 = run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "A1:D1",
            "--font",
            '{"bold": true, "size": 12}',
            "--fill",
            '{"fgColor": "CCCCCC"}',
            "--alignment",
            '{"horizontal": "center"}',
        )
        assert exit1 == 0, f"Format range failed: {out1}"

        # Step 2: Set column widths
        output2 = tmp_path / "step2.xlsx"
        out2, exit2 = run_tool(
            "xls_set_column_width",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--columns",
            "A:D",
            "--width",
            "15",
        )
        assert exit2 == 0, f"Set column width failed: {out2}"

        # Step 3: Freeze panes
        output3 = tmp_path / "step3.xlsx"
        out3, exit3 = run_tool(
            "xls_freeze_panes",
            "--input",
            str(output2),
            "--output",
            str(output3),
            "--freeze",
            "A2",
        )
        assert exit3 == 0, f"Freeze panes failed: {out3}"

        # Step 4: Apply conditional formatting
        output4 = tmp_path / "step4.xlsx"
        out4, exit4 = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(output3),
            "--output",
            str(output4),
            "--range",
            "B2:D10",
            "--type",
            "colorscale",
            "--config",
            '{"start_type": "min", "start_color": "FFFFFF", "end_type": "max", "end_color": "00FF00"}',
        )
        assert exit4 == 0, f"Conditional formatting failed: {out4}"

        # Step 5: Apply number format
        output5 = tmp_path / "final.xlsx"
        out5, exit5 = run_tool(
            "xls_set_number_format",
            "--input",
            str(output4),
            "--output",
            str(output5),
            "--range",
            "B2:D10",
            "--number-format",
            '"$"#,##0.00',
        )
        assert exit5 == 0, f"Number format failed: {out5}"

        # Verify final file has all formatting
        wb = load_workbook(str(output5))
        ws = wb.active

        # Verify header formatting
        assert ws["A1"].font.bold is True
        assert ws["A1"].fill.start_color.rgb == "00CCCCCC"

        # Verify column width
        assert ws.column_dimensions["A"].width == 15.0

        # Verify freeze panes
        assert ws.freeze_panes is not None

        # Verify conditional formatting
        assert len(ws.conditional_formatting._cf_rules) == 1

        # Verify number format
        assert ws["B2"].number_format == '"$"#,##0.00'

    def test_formatting_persistence(self, data_workbook: Path, tmp_path: Path):
        """Test that formatting persists after save/reload."""
        # Apply formatting
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "A1",
            "--font",
            '{"bold": true, "size": 16}',
        )

        # Reload and verify
        wb = load_workbook(str(output1))
        ws = wb.active
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 16

        # Save and reload again
        output2 = tmp_path / "step2.xlsx"
        wb.save(str(output2))

        wb2 = load_workbook(str(output2))
        ws2 = wb2.active
        assert ws2["A1"].font.bold is True
        assert ws2["A1"].font.size == 16


class TestFormattingInteractions:
    """Test interactions between formatting operations."""

    def test_format_then_conditional_format(self, data_workbook: Path, tmp_path: Path):
        """Test applying formatting then conditional formatting."""
        # First format
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_format_range",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "B2:D10",
            "--font",
            '{"size": 10}',
        )

        # Then conditional format
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--range",
            "B2:D10",
            "--type",
            "databar",
            "--config",
            '{"color": "638EC6"}',
        )

        assert exit_code == 0

        # Verify both exist
        wb = load_workbook(str(output2))
        ws = wb.active
        assert ws["B2"].font.size == 10
        assert len(ws.conditional_formatting._cf_rules) == 1

    def test_multiple_number_formats(self, data_workbook: Path, tmp_path: Path):
        """Test applying different number formats to different ranges."""
        # Format column B as currency
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output1),
            "--range",
            "B2:B10",
            "--number-format",
            '"$"#,##0.00',
        )

        # Format column C as percentage
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--range",
            "C2:C10",
            "--number-format",
            "0.00%",
        )

        assert exit_code == 0

        # Verify different formats
        wb = load_workbook(str(output2))
        ws = wb.active
        assert ws["B2"].number_format == '"$"#,##0.00'
        assert ws["C2"].number_format == "0.00%"


class TestFormattingEdgeCases:
    """Test edge cases in formatting operations."""

    def test_empty_range_formatting(self, data_workbook: Path, tmp_path: Path):
        """Test formatting empty cells."""
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_set_number_format",
            "--input",
            str(data_workbook),
            "--output",
            str(output_path),
            "--range",
            "Z1:Z10",  # Empty range
            "--number-format",
            '"$"#,##0.00',
        )

        assert exit_code == 0

    def test_large_range_warning(self, data_workbook: Path, tmp_path: Path):
        """Test performance warning on large ranges."""
        # Create workbook with larger range
        wb = load_workbook(str(data_workbook))
        ws = wb.active

        # Expand to create larger range
        for row in range(1, 101):
            for col in range(1, 101):
                ws.cell(row=row, column=col, value=row * col)

        large_path = tmp_path / "large.xlsx"
        wb.save(str(large_path))

        # Format large range (>10k cells)
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(large_path),
            "--output",
            str(output_path),
            "--range",
            "A1:CX100",  # 10,000+ cells
            "--font",
            '{"bold": true}',
        )

        assert exit_code == 0
        # Should have performance warning
        assert "may be slow" in str(output.get("warnings", [])).lower()


class TestFormattingPerformance:
    """Test performance of formatting operations."""

    @pytest.mark.slow
    def test_format_many_cells(self, data_workbook: Path, tmp_path: Path):
        """Test formatting many cells."""
        # Create larger workbook
        wb = Workbook()
        ws = wb.active
        assert ws is not None

        for row in range(1, 101):
            for col in range(1, 11):
                ws.cell(row=row, column=col, value=f"Cell {row},{col}")

        path = tmp_path / "many_cells.xlsx"
        wb.save(str(path))

        # Format range
        output_path = tmp_path / "output.xlsx"
        output, exit_code = run_tool(
            "xls_format_range",
            "--input",
            str(path),
            "--output",
            str(output_path),
            "--range",
            "A1:J100",  # 1,000 cells
            "--font",
            '{"bold": true}',
            "--fill",
            '{"fgColor": "EEEEEE"}',
        )

        assert exit_code == 0
        assert output["data"]["cells_formatted"] == 1000

    def test_multiple_cf_rules(self, data_workbook: Path, tmp_path: Path):
        """Test applying multiple conditional formatting rules."""
        current_path = data_workbook

        # Add first rule
        output1 = tmp_path / "step1.xlsx"
        run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(current_path),
            "--output",
            str(output1),
            "--range",
            "B2:D10",
            "--type",
            "colorscale",
            "--config",
            '{"start_type": "min", "start_color": "FFFFFF", "end_type": "max", "end_color": "00FF00"}',
        )

        # Add second rule (different range)
        output2 = tmp_path / "step2.xlsx"
        output, exit_code = run_tool(
            "xls_apply_conditional_formatting",
            "--input",
            str(output1),
            "--output",
            str(output2),
            "--range",
            "A2:A10",
            "--type",
            "databar",
            "--config",
            '{"color": "FF0000"}',
        )

        assert exit_code == 0

        # Verify multiple rules
        wb = load_workbook(str(output2))
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) == 2
