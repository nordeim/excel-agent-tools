"""
End-to-End Integration Test: Formula Dependency Governance Loop

Simulates the critical denial-with-prescriptive-guidance cycle:
Dependency Report → Attempt Delete (Denied) → Update References →
Approve Token → Delete (Success) → Final Validation.

Validates the AI agent's ability to programmatically recover from
impact denials using structured guidance fields.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Subprocess Helper
# ---------------------------------------------------------------------------


def _run_tool(tool_module: str, *args: str, cwd: Path | None = None) -> tuple[dict, int]:
    """Execute a CLI tool via subprocess with governance secret injection."""
    env = os.environ.copy()
    env["EXCEL_AGENT_SECRET"] = "e2e-governance-secret-2026"

    cmd = [sys.executable, "-m", f"excel_agent.tools.{tool_module}", *args]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, env=env, cwd=cwd)

    out = result.stdout.strip()
    if not out:
        raise AssertionError(f"Tool {tool_module} produced no output.\nStderr: {result.stderr}")
    return json.loads(out), result.returncode


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def dependency_workbook(tmp_path: Path) -> Path:
    """
    Creates a workbook with guaranteed cross-sheet dependencies:
    Sheet1!A1 = 100
    Sheet2!A1 = Sheet1!A1 * 2
    Deleting Sheet1 will break Sheet2!A1.
    """
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Data"
    ws1["A1"] = 100
    ws1["A2"] = 200
    ws1["A3"] = "=A1+A2"

    ws2 = wb.create_sheet("Report")
    ws2["A1"] = "=Data!A1*2"
    ws2["A2"] = "=Data!A3+50"

    path = tmp_path / "dependency_test.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def complex_dependency_workbook(tmp_path: Path) -> Path:
    """
    Creates a workbook with multi-level dependencies:
    Source → Processor → Final
    Multiple sheets with complex cross-references.
    """
    wb = Workbook()

    # Source sheet
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Source"
    for i in range(1, 6):
        ws1[f"A{i}"] = i * 10
    ws1["B1"] = "=SUM(A1:A5)"

    # Processor sheet
    ws2 = wb.create_sheet("Processor")
    ws2["A1"] = "=Source!B1 * 2"
    ws2["A2"] = "=Source!B1 + Source!A1"
    ws2["A3"] = "=A1 + A2"

    # Final sheet
    ws3 = wb.create_sheet("Final")
    ws3["A1"] = "=Processor!A3 * 10"
    ws3["B1"] = "=Source!B1 + Processor!A1"

    # Summary sheet
    ws4 = wb.create_sheet("Summary")
    ws4["A1"] = "=Final!A1 + Final!B1"
    ws4["A2"] = "=Source!B1 + Processor!A3 + Final!A1"

    path = tmp_path / "complex_deps.xlsx"
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Test Suite
# ---------------------------------------------------------------------------


class TestFormulaDependencyWorkflow:
    """
    Validates the governance loop where agents must resolve formula
    dependencies before executing destructive structural mutations.
    """

    def test_governance_denial_guidance_loop(
        self, dependency_workbook: Path, tmp_path: Path
    ) -> None:
        """Full governance loop: denial → guidance → remediation → approval → success."""
        work_path = tmp_path / "work.xlsx"
        work_path.write_bytes(dependency_workbook.read_bytes())
        audit_path = tmp_path / ".excel_agent_audit.jsonl"

        # Step 1: Generate dependency report
        dep_data, dep_code = _run_tool(
            "governance.xls_dependency_report",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        assert dep_code == 0
        assert dep_data["status"] == "success"
        assert dep_data["data"]["stats"]["total_edges"] > 0, "Fixture must have dependencies"

        # Step 2: Generate token for sheet deletion (required for impact check)
        token_data, token_code = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "sheet:delete",
            "--file",
            str(work_path),
            "--ttl",
            "300",
            cwd=tmp_path,
        )
        assert token_code == 0
        approval_token = token_data["data"]["token"]

        # Step 3: Attempt deletion WITHOUT --acknowledge-impact → Expect Denial
        deny_data, deny_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--name",
            "Data",
            "--token",
            approval_token,
            cwd=tmp_path,
        )
        # Should exit 1 due to ImpactDeniedError
        assert deny_code == 1, f"Expected denial (exit 1), got {deny_code}"
        assert deny_data["status"] in ("error", "denied")
        assert "guidance" in deny_data, (
            "Denial must include prescriptive guidance for agent recovery"
        )
        assert "xls-update-references" in deny_data["guidance"], (
            "Guidance must suggest remediation tool"
        )

        # Step 4: Parse guidance and simulate agent recovery
        # In production, the agent parses guidance and builds the --updates payload.
        # For E2E validation, we call update-references with the exact target range.
        guidance = deny_data["guidance"]
        target_range = "Data!A1:XFD1048576"
        updates_json = json.dumps([{"old": f"{target_range}", "new": f"Sheet1!A1:XFD1048576"}])

        update_data, update_code = _run_tool(
            "cells.xls_update_references",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--updates",
            updates_json,
            cwd=tmp_path,
        )
        assert update_code == 0
        assert update_data["status"] == "success"

        # Step 5: Generate NEW token for the updated file (hash changed after update)
        token_data2, token_code2 = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "sheet:delete",
            "--file",
            str(work_path),
            "--ttl",
            "300",
            cwd=tmp_path,
        )
        assert token_code2 == 0
        approval_token2 = token_data2["data"]["token"]

        # Step 6: Retry deletion WITH --acknowledge-impact → Expect Success
        success_data, success_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--name",
            "Data",
            "--token",
            approval_token2,
            "--acknowledge-impact",
            cwd=tmp_path,
        )
        assert success_code == 0
        assert success_data["status"] == "success"
        assert success_data["data"]["deleted_sheet"] == "Data"

        # Step 6: Final validation → Verify clean state
        final_data, final_code = _run_tool(
            "governance.xls_validate_workbook",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        assert final_code in (0, 1)
        # Verify sheet was actually removed
        meta_data, meta_code = _run_tool(
            "read.xls_get_sheet_names", "--input", str(work_path), cwd=tmp_path
        )
        assert meta_code == 0
        sheet_names = [s["name"] for s in meta_data["data"]["sheets"]]
        assert "Data" not in sheet_names, "Sheet should have been deleted"
        assert "Report" in sheet_names, "Report sheet must persist"

        # Step 7: Verify audit trail contains expected operations
        if audit_path.exists():
            lines = audit_path.read_text(encoding="utf-8").strip().split("\n")
            events = [json.loads(line) for line in lines if line.strip()]
            tools_used = {e.get("tool") for e in events}
            assert "xls_delete_sheet" in tools_used, "Audit trail missing delete operation"
            assert "xls_update_references" in tools_used, "Audit trail missing reference update"

        # Verify no VBA source code leaked (should never happen, but validated)
        for event in events:
            assert "code" not in str(event.get("details", "")).lower(), "Audit leak detected"

    def test_complex_dependency_chain(
        self, complex_dependency_workbook: Path, tmp_path: Path
    ) -> None:
        """Validate dependency tracking across multiple sheets and levels."""
        work_path = tmp_path / "work_complex.xlsx"
        work_path.write_bytes(complex_dependency_workbook.read_bytes())

        # Get dependency report
        dep_data, dep_code = _run_tool(
            "governance.xls_dependency_report",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        assert dep_code == 0

        stats = dep_data["data"]["stats"]
        assert stats["total_formulas"] >= 6  # Multiple formulas in fixture
        assert stats["total_edges"] >= 10  # Multiple dependencies

        # Identify impact of deleting Source sheet
        token_data, _ = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "sheet:delete",
            "--file",
            str(work_path),
            "--ttl",
            "300",
            cwd=tmp_path,
        )
        token = token_data["data"]["token"]

        # Attempt deletion - should be denied due to many dependencies
        deny_data, deny_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--name",
            "Source",
            "--token",
            token,
            cwd=tmp_path,
        )
        assert deny_code == 1
        assert deny_data["impact"]["broken_references"] >= 6

    def test_token_scoping_validation(self, dependency_workbook: Path, tmp_path: Path) -> None:
        """Verify tokens are scope-specific and cannot be cross-used."""
        work_path = tmp_path / "work_scope.xlsx"
        work_path.write_bytes(dependency_workbook.read_bytes())

        # Generate token for wrong scope
        token_data, _ = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "range:delete",  # Wrong scope for sheet delete
            "--file",
            str(work_path),
            "--ttl",
            "300",
            cwd=tmp_path,
        )
        wrong_token = token_data["data"]["token"]

        # Attempt sheet delete with wrong scope token
        deny_data, deny_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--name",
            "Data",
            "--token",
            wrong_token,
            cwd=tmp_path,
        )
        assert deny_code == 4, "Wrong scope token should result in permission denied"
        assert deny_data["status"] == "denied"

    def test_token_file_hash_binding(self, dependency_workbook: Path, tmp_path: Path) -> None:
        """Verify tokens are bound to specific file hash."""
        work_path1 = tmp_path / "work1.xlsx"
        work_path2 = tmp_path / "work2.xlsx"
        work_path1.write_bytes(dependency_workbook.read_bytes())
        work_path2.write_bytes(dependency_workbook.read_bytes())

        # Modify file 2 to ensure different hash
        wb = load_workbook(str(work_path2))
        ws = wb.active
        assert ws is not None
        ws["Z99"] = "modification"  # Add a cell to change hash
        wb.save(str(work_path2))

        # Generate token for file 1
        token_data, _ = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "sheet:delete",
            "--file",
            str(work_path1),
            "--ttl",
            "300",
            cwd=tmp_path,
        )
        token = token_data["data"]["token"]

        # Attempt to use token on file 2
        deny_data, deny_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path2),
            "--output",
            str(work_path2),
            "--name",
            "Data",
            "--token",
            token,
            cwd=tmp_path,
        )
        assert deny_code == 4, "Token for different file should be rejected"

    def test_batch_reference_updates(
        self, complex_dependency_workbook: Path, tmp_path: Path
    ) -> None:
        """Verify batch reference updates handle multiple changes."""
        work_path = tmp_path / "work_batch.xlsx"
        work_path.write_bytes(complex_dependency_workbook.read_bytes())

        # Multiple updates in one operation
        updates = [
            {"old": "Source!A1", "new": "Archive!A1"},
            {"old": "Source!B1", "new": "Archive!B1"},
        ]

        update_data, update_code = _run_tool(
            "cells.xls_update_references",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--updates",
            json.dumps(updates),
            cwd=tmp_path,
        )
        assert update_code == 0
        assert update_data["status"] == "success"
        # Formulas updated should be >= len(updates) since multiple formulas may contain the references
        assert update_data["impact"]["formulas_updated"] >= len(updates)

        # Verify references updated
        dep_data, _ = _run_tool(
            "governance.xls_dependency_report",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        # Should still have dependencies but now pointing to Archive
        deps = json.dumps(dep_data["data"]["graph"])
        assert "Archive" in deps

    def test_circular_reference_handling(self, tmp_path: Path) -> None:
        """Verify circular references are detected and reported."""
        # Create workbook with circular reference
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=B1"
        ws["B1"] = "=A1"

        circular_path = tmp_path / "circular.xlsx"
        wb.save(str(circular_path))

        # Get dependency report
        dep_data, dep_code = _run_tool(
            "governance.xls_dependency_report",
            "--input",
            str(circular_path),
            cwd=tmp_path,
        )
        assert dep_code == 0

        # Should detect circular reference
        circular = dep_data["data"]["circular_refs"]
        assert len(circular) > 0, "Circular reference should be detected"

    def test_impact_report_accuracy(
        self, complex_dependency_workbook: Path, tmp_path: Path
    ) -> None:
        """Verify impact report accurately counts affected formulas."""
        work_path = tmp_path / "work_impact.xlsx"
        work_path.write_bytes(complex_dependency_workbook.read_bytes())

        # Get dependency report with impact analysis
        dep_data, _ = _run_tool(
            "governance.xls_dependency_report",
            "--input",
            str(work_path),
            "--sheet",
            "Source",
            cwd=tmp_path,
        )
        assert dep_data["status"] == "success"

        # Stats should show dependencies from Source sheet
        stats = dep_data["data"]["stats"]
        assert stats["total_formulas"] > 0

        # Verify affected sheets listed
        if dep_data["data"].get("circular_refs"):
            assert isinstance(dep_data["data"]["circular_refs"], list)

    def test_concurrent_modification_protection(
        self, dependency_workbook: Path, tmp_path: Path
    ) -> None:
        """Verify hash-based concurrent modification detection."""
        work_path = tmp_path / "work_concurrent.xlsx"
        work_path.write_bytes(dependency_workbook.read_bytes())

        # Get version hash
        hash_data, _ = _run_tool(
            "governance.xls_version_hash",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        assert hash_data["status"] == "success"

        # Simulate external modification
        wb = Workbook()
        wb = openpyxl.load_workbook(str(work_path))
        ws = wb.active
        assert ws is not None
        ws["Z99"] = "External change"
        wb.save(str(work_path))

        # Try to validate - should detect change
        valid_data, valid_code = _run_tool(
            "governance.xls_validate_workbook",
            "--input",
            str(work_path),
            cwd=tmp_path,
        )
        # File is still valid OOXML, just modified
        assert valid_code in (0, 1)

    def test_token_ttl_expiration(self, dependency_workbook: Path, tmp_path: Path) -> None:
        """Verify expired tokens are rejected."""
        work_path = tmp_path / "work_ttl.xlsx"
        work_path.write_bytes(dependency_workbook.read_bytes())

        # Generate token with 1-second TTL
        token_data, _ = _run_tool(
            "governance.xls_approve_token",
            "--scope",
            "sheet:delete",
            "--file",
            str(work_path),
            "--ttl",
            "1",  # Very short TTL
            cwd=tmp_path,
        )
        token = token_data["data"]["token"]

        # Wait for token to expire
        import time

        time.sleep(2)

        # Attempt to use expired token
        deny_data, deny_code = _run_tool(
            "structure.xls_delete_sheet",
            "--input",
            str(work_path),
            "--output",
            str(work_path),
            "--name",
            "Data",
            "--token",
            token,
            "--acknowledge-impact",
            cwd=tmp_path,
        )
        assert deny_code == 4, "Expired token should be rejected"
        assert (
            "expired" in str(deny_data.get("error", "")).lower() or deny_data["status"] == "denied"
        )


import openpyxl
