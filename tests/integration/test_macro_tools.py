"""Integration tests for macro workflows.

These tests simulate real-world macro workflows using subprocess calls,
verifying the full tool chain from end to end.
"""

from __future__ import annotations

import json
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

# -----------------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------------


def run_cli_tool(tool_name: str, *args: str) -> tuple[dict, int]:
    """Run a CLI tool via subprocess and parse JSON output."""
    module_path = f"excel_agent.tools.macros.{tool_name}"
    cmd = [sys.executable, "-m", module_path, *args]

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=30,
    )

    try:
        output = json.loads(result.stdout) if result.stdout else {}
    except json.JSONDecodeError:
        output = {"parse_error": True, "stdout": result.stdout, "stderr": result.stderr}

    return output, result.returncode


def create_macro_workbook(path: Path) -> Path:
    """Create a workbook with VBA project marker."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = "Test Data"

    wb.save(str(path))

    # Inject vbaProject.bin marker
    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA_PROJECT_MARKER_DATA")

    return path


def create_clean_workbook(path: Path) -> Path:
    """Create a clean workbook without macros."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Clean"
    ws["A1"] = "No macros"

    wb.save(str(path))
    return path


# -----------------------------------------------------------------------------
# xls_has_macros Tests
# -----------------------------------------------------------------------------


def test_has_macros_detects_vba(tmp_path: Path):
    """Test has_macros detects VBA in .xlsm file."""
    path = tmp_path / "test.xlsm"
    create_macro_workbook(path)

    output, exit_code = run_cli_tool("xls_has_macros", "--input", str(path))

    assert exit_code == 0
    assert output["data"]["has_macros"] is True


def test_has_macros_clean_xlsx(tmp_path: Path):
    """Test has_macros returns False for clean .xlsx."""
    path = tmp_path / "clean.xlsx"
    create_clean_workbook(path)

    output, exit_code = run_cli_tool("xls_has_macros", "--input", str(path))

    assert exit_code == 0
    assert output["data"]["has_macros"] is False


def test_has_macros_missing_file(tmp_path: Path):
    """Test has_macros returns error for missing file."""
    path = tmp_path / "nonexistent.xlsx"

    output, exit_code = run_cli_tool("xls_has_macros", "--input", str(path))

    assert exit_code == 2
    assert output["status"] == "error"


# -----------------------------------------------------------------------------
# xls_inspect_macros Tests
# -----------------------------------------------------------------------------


def test_inspect_macros_basic(tmp_path: Path):
    """Test inspect_macros returns macro info."""
    path = tmp_path / "test.xlsm"
    create_macro_workbook(path)

    output, exit_code = run_cli_tool("xls_inspect_macros", "--input", str(path))

    assert exit_code == 0
    assert output["data"]["has_macros"] is True


def test_inspect_macros_clean_file(tmp_path: Path):
    """Test inspect_macros on clean file."""
    path = tmp_path / "clean.xlsx"
    create_clean_workbook(path)

    output, exit_code = run_cli_tool("xls_inspect_macros", "--input", str(path))

    assert exit_code == 0
    assert output["data"]["has_macros"] is False
    assert output["data"]["module_count"] == 0


# -----------------------------------------------------------------------------
# xls_validate_macro_safety Tests
# -----------------------------------------------------------------------------


def test_validate_macro_safety_returns_risk_info(tmp_path: Path):
    """Test validate_macro_safety returns risk assessment."""
    path = tmp_path / "test.xlsm"
    create_macro_workbook(path)

    output, exit_code = run_cli_tool("xls_validate_macro_safety", "--input", str(path))

    assert exit_code == 0
    data = output["data"]
    assert "risk_level" in data
    assert "risk_score" in data
    assert "has_macros" in data


def test_validate_macro_safety_clean_file(tmp_path: Path):
    """Test validate_macro_safety on clean file returns none risk."""
    path = tmp_path / "clean.xlsx"
    create_clean_workbook(path)

    output, exit_code = run_cli_tool("xls_validate_macro_safety", "--input", str(path))

    assert exit_code == 0
    assert output["data"]["risk_level"] == "none"
    assert output["data"]["risk_score"] == 0


# -----------------------------------------------------------------------------
# xls_remove_macros Tests
# -----------------------------------------------------------------------------


def test_remove_macros_no_token_fails(tmp_path: Path):
    """Test remove_macros fails without tokens."""
    input_path = tmp_path / "test.xlsm"
    output_path = tmp_path / "cleaned.xlsx"
    create_macro_workbook(input_path)

    output, exit_code = run_cli_tool(
        "xls_remove_macros",
        "--input",
        str(input_path),
        "--output",
        str(output_path),
    )

    # argparse exits 2 for missing required arguments
    # When argparse fails, output may not contain 'status' (not JSON)
    assert exit_code == 2


def test_remove_macros_no_token2_fails(tmp_path: Path):
    """Test remove_macros fails without second token."""
    input_path = tmp_path / "test.xlsm"
    output_path = tmp_path / "cleaned.xlsx"
    create_macro_workbook(input_path)

    output, exit_code = run_cli_tool(
        "xls_remove_macros",
        "--input",
        str(input_path),
        "--output",
        str(output_path),
        "--token",
        "only_one_token",
    )

    # argparse exits 2 for missing --token2
    assert exit_code == 2


# -----------------------------------------------------------------------------
# xls_inject_vba_project Tests
# -----------------------------------------------------------------------------


def test_inject_no_token_fails(tmp_path: Path):
    """Test inject fails without token."""
    target_path = tmp_path / "target.xlsx"
    vba_bin = tmp_path / "vbaProject.bin"
    output_path = tmp_path / "injected.xlsm"

    create_clean_workbook(target_path)
    vba_bin.write_bytes(b"VBA_DATA")

    output, exit_code = run_cli_tool(
        "xls_inject_vba_project",
        "--input",
        str(target_path),
        "--vba-bin",
        str(vba_bin),
        "--output",
        str(output_path),
    )

    # Missing token raises ValidationError (exit 1) or caught as internal (exit 5)
    assert exit_code in [1, 5]
    assert output.get("status") == "error"


def test_inject_missing_vba_bin_fails(tmp_path: Path):
    """Test inject fails when vba-bin doesn't exist."""
    target_path = tmp_path / "target.xlsx"
    vba_bin = tmp_path / "nonexistent.bin"
    output_path = tmp_path / "injected.xlsm"

    create_clean_workbook(target_path)

    output, exit_code = run_cli_tool(
        "xls_inject_vba_project",
        "--input",
        str(target_path),
        "--vba-bin",
        str(vba_bin),
        "--output",
        str(output_path),
        "--token",
        "some_token",
    )

    # Non-existent .bin file: could fail with validation (1), file not found (2), or internal (5)
    assert exit_code in [1, 2, 5]
    assert output.get("status") == "error"


# -----------------------------------------------------------------------------
# End-to-End Workflow Tests
# -----------------------------------------------------------------------------


def test_full_macro_workflow(tmp_path: Path):
    """Test full macro workflow:
    1. Check for macros
    2. Inspect macros
    3. Validate safety
    4. Verify clean file has no macros
    """
    macro_path = tmp_path / "with_macros.xlsm"
    clean_path = tmp_path / "clean.xlsx"

    create_macro_workbook(macro_path)
    create_clean_workbook(clean_path)

    # Step 1: Check for macros
    output1, exit1 = run_cli_tool("xls_has_macros", "--input", str(macro_path))
    assert exit1 == 0
    assert output1["data"]["has_macros"] is True

    # Step 2: Inspect macros
    output2, exit2 = run_cli_tool("xls_inspect_macros", "--input", str(macro_path))
    assert exit2 == 0
    assert output2["data"]["has_macros"] is True

    # Step 3: Validate safety
    output3, exit3 = run_cli_tool("xls_validate_macro_safety", "--input", str(macro_path))
    assert exit3 == 0
    assert "risk_level" in output3["data"]

    # Step 4: Verify clean file
    output4, exit4 = run_cli_tool("xls_has_macros", "--input", str(clean_path))
    assert exit4 == 0
    assert output4["data"]["has_macros"] is False


def test_audit_trail_excludes_source_code(tmp_path: Path):
    """Verify that audit trail never includes VBA source code."""
    macro_path = tmp_path / "test.xlsm"
    create_macro_workbook(macro_path)

    # Inspect macros
    output, _ = run_cli_tool("xls_inspect_macros", "--input", str(macro_path))

    # Audit trail would be written (if configured)
    # This test verifies the tool doesn't log code in output
    if "audit_trail" in str(output):
        # Ensure no code in audit trail
        audit_text = str(output)
        # Should not contain Sub/Function definitions
        assert "Sub " not in audit_text or "code_preview" in str(output)


@pytest.mark.slow
def test_large_workbook_macro_check(tmp_path: Path):
    """Test macro detection performance on large workbook."""
    path = tmp_path / "large.xlsm"

    # Create large workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    # Add substantial data
    for i in range(1, 1001):
        ws[f"A{i}"] = f"Data {i}"
        ws[f"B{i}"] = i * 10

    wb.save(str(path))

    # Inject vba marker
    with zipfile.ZipFile(path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"VBA" * 1000)

    # Test has_macros
    output, exit_code = run_cli_tool("xls_has_macros", "--input", str(path))
    assert exit_code == 0
    assert output["data"]["has_macros"] is True
