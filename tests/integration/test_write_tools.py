"""Integration tests for write and create tools via subprocess."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def _run_tool(tool_module: str, *args: str) -> dict:
    """Run a CLI tool via subprocess and return parsed JSON output."""
    result = subprocess.run(  # noqa: S603
        [sys.executable, "-m", f"excel_agent.tools.{tool_module}", *args],
        capture_output=True,
        text=True,
        timeout=30,
    )
    assert result.stdout.strip(), f"Tool {tool_module} produced no output. stderr: {result.stderr}"
    return json.loads(result.stdout)


class TestCreateNew:
    """Tests for xls_create_new."""

    def test_create_default(self, tmp_path: Path) -> None:
        output = tmp_path / "new.xlsx"
        result = _run_tool("write.xls_create_new", "--output", str(output))
        assert result["status"] == "success"
        assert output.exists()
        wb = load_workbook(str(output))
        assert wb.sheetnames == ["Sheet1"]

    def test_create_with_sheets(self, tmp_path: Path) -> None:
        output = tmp_path / "multi.xlsx"
        result = _run_tool(
            "write.xls_create_new",
            "--output",
            str(output),
            "--sheets",
            "Data,Summary,Charts",
        )
        assert result["status"] == "success"
        wb = load_workbook(str(output))
        assert wb.sheetnames == ["Data", "Summary", "Charts"]

    def test_create_single_named(self, tmp_path: Path) -> None:
        output = tmp_path / "single.xlsx"
        result = _run_tool(
            "write.xls_create_new",
            "--output",
            str(output),
            "--sheets",
            "MyData",
        )
        assert result["status"] == "success"
        wb = load_workbook(str(output))
        assert wb.sheetnames == ["MyData"]


class TestCreateFromTemplate:
    """Tests for xls_create_from_template."""

    def test_variable_substitution(self, tmp_path: Path) -> None:
        # Create a template with placeholders
        from openpyxl import Workbook

        template = tmp_path / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "{{company}}"
        ws["A2"] = "Report for {{year}}"
        ws["A3"] = "=SUM(B1:B10)"  # Formula should NOT be substituted
        wb.save(str(template))

        output = tmp_path / "from_template.xlsx"
        result = _run_tool(
            "write.xls_create_from_template",
            "--template",
            str(template),
            "--output",
            str(output),
            "--vars",
            '{"company": "Acme Corp", "year": "2026"}',
        )
        assert result["status"] == "success"
        assert result["data"]["substitutions_made"] == 2

        wb2 = load_workbook(str(output))
        ws2 = wb2.active
        assert ws2 is not None
        assert ws2["A1"].value == "Acme Corp"
        assert ws2["A2"].value == "Report for 2026"
        # Formula must be preserved, not substituted
        assert ws2["A3"].value == "=SUM(B1:B10)"

    def test_unmatched_placeholders_preserved(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        template = tmp_path / "template2.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "{{known}}"
        ws["A2"] = "{{unknown}}"
        wb.save(str(template))

        output = tmp_path / "partial.xlsx"
        _run_tool(
            "write.xls_create_from_template",
            "--template",
            str(template),
            "--output",
            str(output),
            "--vars",
            '{"known": "replaced"}',
        )
        wb2 = load_workbook(str(output))
        ws2 = wb2.active
        assert ws2 is not None
        assert ws2["A1"].value == "replaced"
        assert ws2["A2"].value == "{{unknown}}"


class TestWriteRange:
    """Tests for xls_write_range."""

    def test_write_basic_data(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_range",
            "--input",
            str(work),
            "--output",
            str(work),
            "--range",
            "F1",
            "--sheet",
            "Sheet1",
            "--data",
            '[["Extra", "Col"], ["X", 42]]',
        )
        assert result["status"] == "success"
        assert result["impact"]["cells_modified"] == 4

        wb = load_workbook(str(work))
        ws = wb["Sheet1"]
        assert ws["F1"].value == "Extra"
        assert ws["G2"].value == 42

    def test_write_formula(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_range",
            "--input",
            str(work),
            "--output",
            str(work),
            "--range",
            "H1",
            "--sheet",
            "Sheet1",
            "--data",
            '[["=A1+1"]]',
        )
        assert result["status"] == "success"
        assert result["impact"]["formulas_updated"] == 1

        wb = load_workbook(str(work))
        ws = wb["Sheet1"]
        assert ws["H1"].data_type == "f"

    def test_write_dates(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_range",
            "--input",
            str(work),
            "--output",
            str(work),
            "--range",
            "H1",
            "--sheet",
            "Sheet1",
            "--data",
            '[["2026-04-08"]]',
        )
        assert result["status"] == "success"
        wb = load_workbook(str(work))
        ws = wb["Sheet1"]
        # Should be stored as datetime, not string
        import datetime

        assert isinstance(ws["H1"].value, (datetime.date, datetime.datetime))


class TestWriteCell:
    """Tests for xls_write_cell."""

    def test_write_with_auto_inference(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_cell",
            "--input",
            str(work),
            "--output",
            str(work),
            "--cell",
            "H1",
            "--sheet",
            "Sheet1",
            "--value",
            "42",
        )
        assert result["status"] == "success"
        wb = load_workbook(str(work))
        assert wb["Sheet1"]["H1"].value == 42

    def test_write_with_explicit_type(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_cell",
            "--input",
            str(work),
            "--output",
            str(work),
            "--cell",
            "H1",
            "--sheet",
            "Sheet1",
            "--value",
            "2026-04-08",
            "--type",
            "date",
        )
        assert result["status"] == "success"
        import datetime

        wb = load_workbook(str(work))
        assert isinstance(wb["Sheet1"]["H1"].value, (datetime.date, datetime.datetime))

    def test_write_formula_explicit(self, sample_workbook: Path, tmp_path: Path) -> None:
        import shutil

        work = tmp_path / "work.xlsx"
        shutil.copy2(sample_workbook, work)

        result = _run_tool(
            "write.xls_write_cell",
            "--input",
            str(work),
            "--output",
            str(work),
            "--cell",
            "H1",
            "--sheet",
            "Sheet1",
            "--value",
            "SUM(A1:A10)",
            "--type",
            "formula",
        )
        assert result["status"] == "success"
        assert result["data"]["is_formula"] is True
        wb = load_workbook(str(work))
        assert wb["Sheet1"]["H1"].value == "=SUM(A1:A10)"
