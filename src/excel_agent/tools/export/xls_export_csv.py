"""xls_export_csv: Export Excel sheet to CSV with encoding control.

Supports multiple encodings (utf-8, latin-1, cp1252, etc.) and delimiters.
Uses streaming to handle large files efficiently.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Supported encodings
SUPPORTED_ENCODINGS = {
    "utf-8": "UTF-8 (default, recommended)",
    "utf-16": "UTF-16 (for wide characters)",
    "latin-1": "Latin-1 (ISO-8859-1, Western European)",
    "cp1252": "Windows-1252 (Western European Windows)",
    "ascii": "ASCII (7-bit, limited characters)",
}

# Supported delimiters
SUPPORTED_DELIMITERS = {
    ",": "comma (default)",
    ";": "semicolon",
    "\t": "tab",
    "|": "pipe",
}


def _run() -> dict[str, object]:
    parser = create_parser("Export Excel sheet to CSV.")
    add_common_args(parser)
    # Note: --output is already added by add_common_args
    parser.add_argument(
        "--encoding",
        type=str,
        default="utf-8",
        help=f"Output encoding (default: utf-8). Supported: {', '.join(SUPPORTED_ENCODINGS.keys())}",
    )
    parser.add_argument(
        "--delimiter",
        type=str,
        default=",",
        help=f"CSV delimiter (default: comma). Supported: {', '.join(SUPPORTED_DELIMITERS.keys())}",
    )
    parser.add_argument(
        "--include-headers",
        action="store_true",
        default=True,
        help="Include header row (default: True)",
    )
    parser.add_argument(
        "--outfile",
        type=str,
        required=False,
        help="Output CSV file path (default: same name with .csv extension)",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(
        args.outfile or str(input_path.with_suffix(".csv")), create_parents=True
    )

    file_hash = compute_file_hash(input_path)

    # Validate encoding
    encoding = args.encoding.lower()
    if encoding not in SUPPORTED_ENCODINGS:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[
                f"Unsupported encoding: {args.encoding}",
                f"Supported: {', '.join(SUPPORTED_ENCODINGS.keys())}",
            ],
        )

    # Validate delimiter
    delimiter = args.delimiter
    if delimiter not in SUPPORTED_DELIMITERS:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[
                f"Unsupported delimiter: {repr(delimiter)}",
                f"Supported: {', '.join(repr(d) for d in SUPPORTED_DELIMITERS.keys())}",
            ],
        )

    # Load workbook in read-only mode for streaming
    try:
        wb = load_workbook(str(input_path), data_only=True, read_only=True)
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to load workbook: {e}"],
        )

    # Select sheet
    try:
        ws = wb[args.sheet] if args.sheet else wb.active
        if ws is None:
            raise ValueError("No active sheet")
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to access sheet: {e}"],
        )

    # Export to CSV
    try:
        with open(output_path, "w", encoding=encoding, newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)

            row_count = 0
            col_count = 0

            for row in ws.iter_rows(values_only=True):
                # Clean None values to empty strings
                cleaned = ["" if v is None else str(v) for v in row]
                writer.writerow(cleaned)
                row_count += 1
                col_count = max(col_count, len(cleaned))

        # Handle empty sheet
        if row_count == 0 and args.include_headers:
            # Create empty CSV with just headers if we can detect them
            pass

    except UnicodeEncodeError as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[
                f"Encoding error: {e}",
                f"Try using --encoding utf-8 or --encoding utf-16 for international characters",
            ],
        )
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=5,
            warnings=[f"Failed to write CSV: {e}"],
        )
    finally:
        wb.close()

    # Get output file info
    output_size = output_path.stat().st_size

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_export_csv",
        scope="read",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "output_path": str(output_path),
            "sheet": ws.title,
            "encoding": encoding,
            "delimiter": repr(delimiter),
            "row_count": row_count,
        },
        impact={"rows_exported": row_count, "columns": col_count},
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "output_path": str(output_path),
            "sheet": ws.title,
            "row_count": row_count,
            "column_count": col_count,
            "encoding": encoding,
            "delimiter": delimiter,
            "file_size_bytes": output_size,
        },
        warnings=None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
