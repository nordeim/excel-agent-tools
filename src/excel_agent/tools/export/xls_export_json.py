"""xls_export_json: Export Excel data to structured JSON.

Supports 3 orientations: records (list of dicts), values (list of lists), columns (dict of lists).
Handles type conversion (dates → ISO strings, None → null).
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Supported orientations
ORIENTATIONS = {
    "records": "List of dictionaries (default)",
    "values": "List of lists (rows)",
    "columns": "Dictionary of column arrays",
}


def _convert_value(value: Any) -> Any:
    """Convert Excel value to JSON-serializable type."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        # Handle NaN, Inf
        if value != value:  # NaN check
            return None
        return value
    return value


def _export_records(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> list[dict]:
    """Export as list of dictionaries."""
    data = []
    headers = None

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=False
        ),
        start=min_row,
    ):
        values = [_convert_value(cell.value) for cell in row]

        if headers is None:
            # First row is headers
            headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(values)]
            continue

        row_dict = {header: val for header, val in zip(headers, values)}
        data.append(row_dict)

    return data


def _export_values(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> list[list]:
    """Export as list of lists (rows)."""
    data = []

    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
    ):
        values = [_convert_value(v) for v in row]
        data.append(values)

    return data


def _export_columns(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> dict[str, list]:
    """Export as dictionary of column arrays."""
    headers = None
    columns: dict[str, list] = {}

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
        ),
        start=min_row,
    ):
        values = [_convert_value(v) for v in row]

        if headers is None:
            headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(values)]
            for header in headers:
                columns[header] = []
            continue

        for header, val in zip(headers, values):
            columns[header].append(val)

    return columns


def _run() -> dict[str, object]:
    parser = create_parser("Export Excel data to JSON.")
    add_common_args(parser)
    parser.add_argument(
        "--outfile",
        type=str,
        required=False,
        help="Output JSON file path (default: same name with .json extension)",
    )
    parser.add_argument(
        "--range",
        type=str,
        default=None,
        help='Range to export (e.g., "A1:D100"). If not specified, exports entire sheet.',
    )
    parser.add_argument(
        "--orient",
        type=str,
        default="records",
        choices=list(ORIENTATIONS.keys()),
        help=f"Output format (default: records). Options: {', '.join(ORIENTATIONS.keys())}",
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        default=False,
        help="Pretty-print JSON with indentation (default: False)",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(
        args.outfile or str(input_path.with_suffix(".json")),
        create_parents=True,
        allowed_suffixes={".json"},
    )

    file_hash = compute_file_hash(input_path)

    # Load workbook in read-only mode
    try:
        wb = load_workbook(str(input_path), data_only=True, read_only=True)
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to load workbook: {e}"],
        )

    # Select sheet
    try:
        ws = wb[args.sheet] if args.sheet else wb.active
        if ws is None:
            raise ValueError("No active sheet")
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to access sheet: {e}"],
        )

    # Determine range
    try:
        if args.range:
            min_col, min_row, max_col, max_row = range_boundaries(args.range)
            if min_col is None or min_row is None:
                raise ValueError("Invalid range")
            # Handle open-ended ranges
            if max_col is None:
                max_col = min_col
            if max_row is None:
                max_row = ws.max_row
        else:
            min_row, min_col = 1, 1
            max_row, max_col = ws.max_row, ws.max_column
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Failed to parse range: {e}"],
        )

    # Export data
    try:
        if args.orient == "records":
            data = _export_records(ws, min_row, max_row, min_col, max_col)
        elif args.orient == "values":
            data = _export_values(ws, min_row, max_row, min_col, max_col)
        elif args.orient == "columns":
            data = _export_columns(ws, min_row, max_row, min_col, max_col)
        else:
            return build_response(
                "error",
                None,
                exit_code=1,
                warnings=[f"Unknown orientation: {args.orient}"],
            )

        # Write JSON
        indent = 2 if args.pretty else None
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=indent, ensure_ascii=False, default=str)

        record_count = (
            len(data)
            if isinstance(data, list)
            else sum(len(v) for v in data.values())
            if isinstance(data, dict)
            else 0
        )

    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=5,
            warnings=[f"Failed to export JSON: {e}"],
        )
    finally:
        wb.close()

    # Get output file info
    output_size = output_path.stat().st_size

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_export_json",
        scope="read",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "output_path": str(output_path),
            "sheet": ws.title,
            "orient": args.orient,
            "range": args.range or "entire_sheet",
            "record_count": record_count,
        },
        impact={"records_exported": record_count},
        success=True,
        exit_code=0,
    )

    # Build preview
    if isinstance(data, list) and data:
        preview = data[:3] if args.orient == "records" else data[:3]
    elif isinstance(data, dict) and data:
        preview = {k: v[:3] for k, v in list(data.items())[:3]}
    else:
        preview = []

    return build_response(
        "success",
        {
            "output_path": str(output_path),
            "sheet": ws.title,
            "orient": args.orient,
            "range": (
                args.range or f"A1:{chr(64 + max_col)}{max_row}" if max_col <= 26 else "full_sheet"
            ),
            "record_count": record_count,
            "file_size_bytes": output_size,
            "preview": preview,
        },
        warnings=None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
