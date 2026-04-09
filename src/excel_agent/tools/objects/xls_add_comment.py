"""xls_add_comment: Add threaded comments to cells.

Creates comments on specific cells with author attribution. Supports
legacy comments (notes) with width/height configuration.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.comments import Comment

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Excel comment limits
MAX_COMMENT_LENGTH = 32767  # Excel's maximum comment length
DEFAULT_WIDTH = 300  # points
DEFAULT_HEIGHT = 100  # points


def _truncate_text(text: str, max_length: int) -> str:
    """Truncate text to max length with indicator."""
    if len(text) <= max_length:
        return text
    return text[: max_length - 3] + "..."


def _run() -> dict[str, object]:
    parser = create_parser("Add comments to cells.")
    add_common_args(parser)
    parser.add_argument(
        "--cell",
        type=str,
        required=True,
        help='Target cell (e.g., "B2")',
    )
    parser.add_argument(
        "--text",
        type=str,
        required=True,
        help="Comment text",
    )
    parser.add_argument(
        "--author",
        type=str,
        default="excel-agent",
        help="Comment author (default: excel-agent)",
    )
    parser.add_argument(
        "--width",
        type=int,
        default=DEFAULT_WIDTH,
        help=f"Comment width in points (default: {DEFAULT_WIDTH})",
    )
    parser.add_argument(
        "--height",
        type=int,
        default=DEFAULT_HEIGHT,
        help=f"Comment height in points (default: {DEFAULT_HEIGHT})",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or str(input_path), create_parents=True)

    file_hash = compute_file_hash(input_path)

    # Validate text
    if not args.text.strip():
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=["Comment text cannot be empty"],
        )

    # Validate cell format
    try:
        from openpyxl.utils import coordinate_to_tuple

        _row, _col = coordinate_to_tuple(args.cell)  # Validate format, values not needed
    except Exception:
        return build_response(
            "error",
            None,
            exit_code=1,
            warnings=[f"Invalid cell reference: {args.cell}"],
        )

    # Load workbook
    wb = load_workbook(str(input_path))
    ws = wb[args.sheet] if args.sheet else wb.active

    # Check if cell has existing comment
    warnings = []
    cell = ws[args.cell]
    if cell.comment:
        warnings.append(f"Cell {args.cell} already has a comment. It will be replaced.")

    # Truncate text if too long
    original_length = len(args.text)
    display_text = args.text
    if original_length > MAX_COMMENT_LENGTH:
        display_text = _truncate_text(args.text, MAX_COMMENT_LENGTH)
        warnings.append(
            f"Comment text truncated from {original_length} to {len(display_text)} characters"
        )

    # Create comment
    try:
        comment = Comment(display_text, args.author)
        comment.width = args.width
        comment.height = args.height
        cell.comment = comment
    except Exception as e:
        return build_response(
            "error",
            None,
            exit_code=5,
            warnings=[f"Failed to create comment: {e}"],
        )

    # Save workbook
    wb.save(str(output_path))

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_add_comment",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "cell": args.cell,
            "author": args.author,
            "text_preview": display_text[:50] + "..." if len(display_text) > 50 else display_text,
            "sheet": ws.title,
        },
        impact={
            "comment_added": True,
            "replaced_existing": cell.comment is not None,
        },
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "cell": args.cell,
            "author": args.author,
            "text_length": len(display_text),
            "sheet": ws.title,
            "width": args.width,
            "height": args.height,
            "truncated": original_length > MAX_COMMENT_LENGTH,
        },
        warnings=warnings if warnings else None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
