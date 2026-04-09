"""xls_create_new: Create a blank workbook with optional sheet names."""

from __future__ import annotations

from openpyxl import Workbook

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import create_parser, validate_output_path
from excel_agent.utils.json_io import build_response


def _run() -> dict:
    parser = create_parser("Create a new blank workbook with specified sheet names.")
    parser.add_argument("--output", type=str, required=True, help="Output workbook path (.xlsx)")
    parser.add_argument(
        "--sheets",
        type=str,
        default=None,
        help="Comma-separated sheet names (default: single 'Sheet1')",
    )
    args = parser.parse_args()

    output = validate_output_path(args.output, create_parents=True)

    wb = Workbook()

    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(",") if s.strip()]
    else:
        sheet_names = ["Sheet1"]

    # A workbook is always created with at least one worksheet.
    # Rename the default sheet to the first name, then create the rest.
    default_ws = wb.active
    assert default_ws is not None

    if sheet_names:
        default_ws.title = sheet_names[0]
        for name in sheet_names[1:]:
            wb.create_sheet(name)
    else:
        default_ws.title = "Sheet1"

    wb.save(str(output))
    file_hash = compute_file_hash(output)

    return build_response(
        "success",
        {
            "output_path": str(output),
            "sheets": list(wb.sheetnames),
            "sheet_count": len(wb.sheetnames),
        },
        workbook_version=file_hash,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
