"""xls_apply_conditional_formatting: Apply conditional formatting rules.

Supports 5 types:
- cellIs: Cell value comparison (greaterThan, lessThan, etc.)
- colorScale: Gradient color scale (2-color or 3-color)
- dataBar: Data bars with color
- iconSet: Icon sets (arrows, traffic lights, etc.)
- formula: Custom formula-based rules
"""

from __future__ import annotations

import json

from openpyxl import load_workbook
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl.styles import PatternFill

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Supported conditional formatting types
CF_TYPES = {
    "cellis": "Cell value comparison",
    "colorscale": "Color scale gradient",
    "databar": "Data bar",
    "iconset": "Icon set",
    "formula": "Custom formula",
}

# Valid operators for cellIs
VALID_OPERATORS = {
    "greaterThan",
    "greaterThanOrEqual",
    "lessThan",
    "lessThanOrEqual",
    "equal",
    "notEqual",
    "between",
    "notBetween",
}

# Icon set styles
ICON_SETS = {
    "3Arrows": "3 arrows (up, sideways, down)",
    "3ArrowsGray": "3 gray arrows",
    "3Flags": "3 flags",
    "3TrafficLights1": "3 traffic lights",
    "3TrafficLights2": "3 traffic lights (rimmed)",
    "3Signs": "3 signs",
    "3Symbols": "3 symbols (circled)",
    "3Symbols2": "3 symbols (uncircled)",
    "4Arrows": "4 arrows",
    "4ArrowsGray": "4 gray arrows",
    "4RedToBlack": "4 red to black",
    "4TrafficLights": "4 traffic lights",
    "5Arrows": "5 arrows",
    "5ArrowsGray": "5 gray arrows",
    "5Rating": "5 rating stars",
    "5Quarters": "5 quarters",
}


def _parse_config(config_str: str) -> dict:
    """Parse JSON config string."""
    try:
        return json.loads(config_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON config: {e}")


def _create_fill(color: str) -> PatternFill:
    """Create solid fill with color."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _create_cellis_rule(config: dict) -> CellIsRule:
    """Create CellIsRule from config."""
    operator = config.get("operator", "greaterThan")
    if operator not in VALID_OPERATORS:
        raise ValueError(f"Invalid operator: {operator}. Valid: {VALID_OPERATORS}")

    formula = config.get("formula", ["0"])
    if not isinstance(formula, list):
        formula = [formula]

    fill_config = config.get("fill", {})
    fill = _create_fill(fill_config.get("fgColor", "FF0000"))

    return CellIsRule(operator=operator, formula=formula, fill=fill)


def _create_colorscale_rule(config: dict) -> ColorScaleRule:
    """Create ColorScaleRule from config."""
    start_type = config.get("start_type", "min")
    start_color = config.get("start_color", "FF0000")
    end_type = config.get("end_type", "max")
    end_color = config.get("end_color", "00FF00")

    # Check for mid point (3-color scale)
    mid_type = config.get("mid_type")
    mid_color = config.get("mid_color")

    if mid_type and mid_color:
        return ColorScaleRule(
            start_type=start_type,
            start_color=start_color,
            mid_type=mid_type,
            mid_color=mid_color,
            end_type=end_type,
            end_color=end_color,
        )
    return ColorScaleRule(
        start_type=start_type,
        start_color=start_color,
        end_type=end_type,
        end_color=end_color,
    )


def _create_databar_rule(config: dict) -> DataBarRule:
    """Create DataBarRule from config."""
    start_type = config.get("start_type", "min")
    end_type = config.get("end_type", "max")
    color = config.get("color", "638EC6")
    show_value = config.get("showValue", True)
    min_length = config.get("min_length")
    max_length = config.get("max_length")

    return DataBarRule(
        start_type=start_type,
        end_type=end_type,
        color=color,
        showValue=show_value,
        minLength=min_length,
        maxLength=max_length,
    )


def _create_iconset_rule(config: dict) -> IconSetRule:
    """Create IconSetRule from config."""
    icon_style = config.get("icon_style", "3Arrows")
    if icon_style not in ICON_SETS:
        raise ValueError(f"Invalid icon_style: {icon_style}. Valid: {list(ICON_SETS.keys())}")

    rule_type = config.get("type", "percent")
    values = config.get("values", [0, 33, 67])
    show_value = config.get("show_value", True)

    return IconSetRule(
        icon_style=icon_style,
        type=rule_type,
        values=values,
        showValue=show_value,
    )


def _create_formula_rule(config: dict) -> FormulaRule:
    """Create FormulaRule from config."""
    formula = config.get("formula", ["TRUE()"])
    if not isinstance(formula, list):
        formula = [formula]

    fill_config = config.get("fill", {})
    fill = _create_fill(fill_config.get("fgColor", "FF0000"))

    return FormulaRule(formula=formula, fill=fill)


def _run() -> dict[str, object]:
    parser = create_parser("Apply conditional formatting to cell ranges.")
    add_common_args(parser)
    parser.add_argument(
        "--range",
        type=str,
        required=True,
        help='Target range (e.g., "A1:A100")',
    )
    parser.add_argument(
        "--type",
        type=str,
        required=True,
        choices=list(CF_TYPES.keys()),
        help=f"Conditional formatting type: {', '.join(CF_TYPES.keys())}",
    )
    parser.add_argument(
        "--config",
        type=str,
        required=True,
        help="JSON configuration for the rule type",
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or str(input_path), create_parents=True)

    file_hash = compute_file_hash(input_path)

    # Parse config
    try:
        config = _parse_config(args.config)
    except ValueError as e:
        return build_response("error", None, exit_code=1, warnings=[str(e)])

    # Load workbook
    wb = load_workbook(str(input_path))
    ws = wb[args.sheet] if args.sheet else wb.active
    if ws is None:
        return build_response("error", None, exit_code=1, warnings=["No active sheet found"])

    # Create appropriate rule
    try:
        cf_type = args.type.lower()
        if cf_type == "cellis":
            rule = _create_cellis_rule(config)
        elif cf_type == "colorscale":
            rule = _create_colorscale_rule(config)
        elif cf_type == "databar":
            rule = _create_databar_rule(config)
        elif cf_type == "iconset":
            rule = _create_iconset_rule(config)
        elif cf_type == "formula":
            rule = _create_formula_rule(config)
        else:
            return build_response(
                "error", None, exit_code=1, warnings=[f"Unknown type: {cf_type}"]
            )
    except ValueError as e:
        return build_response("error", None, exit_code=1, warnings=[str(e)])
    except Exception as e:
        return build_response("error", None, exit_code=5, warnings=[f"Failed to create rule: {e}"])

    # Add rule to worksheet
    try:
        ws.conditional_formatting.add(args.range, rule)
    except Exception as e:
        return build_response(
            "error", None, exit_code=5, warnings=[f"Failed to apply formatting: {e}"]
        )

    # Save workbook
    wb.save(str(output_path))

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_apply_conditional_formatting",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "range": args.range,
            "type": args.type,
            "config": config,
            "sheet": ws.title,
        },
        impact={"rule_added": True, "cf_type": args.type},
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "range": args.range,
            "type": args.type,
            "config": config,
            "sheet": ws.title,
        },
        warnings=None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
