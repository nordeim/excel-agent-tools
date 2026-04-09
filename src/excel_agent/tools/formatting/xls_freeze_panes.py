"""xls_freeze_panes: Freeze rows and columns for scrolling.

Freezes panes at specified cell position. Common patterns:
- 'A2': Freeze first row
- 'B1': Freeze first column
- 'B2': Freeze first row and first column
- 'none': Unfreeze all panes
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response


def _parse_freeze_position(freeze_str: str) -> tuple[int, int] | None:
    """Parse freeze position string to (row, col) tuple.

    Returns None for 'none' (unfreeze).
    """
    freeze_lower = freeze_str.lower().strip()

    if freeze_lower in ("none", "unfreeze", "clear"):
        return None

    # Validate cell reference
    try:
        row, col = coordinate_to_tuple(freeze_str)
        return (row, col)
    except Exception as e:
        raise ValueError(f"Invalid freeze position '{freeze_str}': {e}")


def _run() -> dict[str, object]:
    parser = create_parser("Freeze rows and columns for scrolling.")
    add_common_args(parser)
    parser.add_argument(
        "--freeze",
        type=str,
        required=True,
        help='Freeze position (e.g., "B2", "A2" freezes row 1, "none" to unfreeze)',
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or str(input_path), create_parents=True)

    file_hash = compute_file_hash(input_path)

    # Parse freeze position
    try:
        position = _parse_freeze_position(args.freeze)
    except ValueError as e:
        return build_response("error", None, exit_code=1, warnings=[str(e)])

    # Load workbook
    wb = load_workbook(str(input_path))
    ws = wb[args.sheet] if args.sheet else wb.active
    if ws is None:
        return build_response("error", None, exit_code=1, warnings=["No active sheet found"])

    # Calculate freeze info
    if position is None:
        # Unfreeze
        ws.freeze_panes = None
        frozen_rows = 0
        frozen_cols = 0
    else:
        row, col = position
        f"{getattr(ws, '_cells', {}).get((row, 1), f'A{row}')}"

        # The freeze position is the top-left cell of the scrollable area
        # So row-1 rows are frozen, col-1 columns are frozen
        frozen_rows = row - 1
        frozen_cols = col - 1

        # Convert to cell reference for openpyxl
        from openpyxl.utils import get_column_letter

        freeze_cell_ref = f"{get_column_letter(col)}{row}"
        ws.freeze_panes = freeze_cell_ref

    # Save workbook
    wb.save(str(output_path))

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_freeze_panes",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "freeze": args.freeze,
            "sheet": ws.title,
            "frozen_rows": frozen_rows,
            "frozen_cols": frozen_cols,
        },
        impact={
            "panes_frozen": position is not None,
            "frozen_rows": frozen_rows,
            "frozen_cols": frozen_cols,
        },
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "freeze": args.freeze,
            "sheet": ws.title,
            "frozen_rows": frozen_rows,
            "frozen_cols": frozen_cols,
            "is_frozen": position is not None,
        },
        warnings=None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
