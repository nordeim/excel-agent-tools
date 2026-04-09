"""xls_set_column_width: Set column widths (fixed or auto-fit).

Supports fixed width values or automatic width calculation based on content.
Auto-fit uses heuristic: max content length + 2 padding, capped at 50.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from excel_agent.core.version_hash import compute_file_hash
from excel_agent.governance.audit_trail import AuditTrail
from excel_agent.tools._tool_base import run_tool
from excel_agent.utils.cli_helpers import (
    add_common_args,
    create_parser,
    validate_input_path,
    validate_output_path,
)
from excel_agent.utils.json_io import build_response

# Excel column width limits
MIN_WIDTH = 0
MAX_WIDTH = 255
DEFAULT_WIDTH = 8.43
AUTO_FIT_CAP = 50
AUTO_FIT_PADDING = 2


def _parse_columns(columns_str: str) -> list[int]:
    """Parse column specification string to list of column indices.

    Supports:
    - Single column: "A"
    - Multiple columns: "A,C,E"
    - Range: "A:C"
    - Mixed: "A,C:E,G"
    """
    columns = set()
    parts = columns_str.split(",")

    for part in parts:
        part = part.strip()
        if ":" in part:
            # Range
            start, end = part.split(":")
            start_col = range_boundaries(f"{start}1")[0]
            end_col = range_boundaries(f"{end}1")[0]
            if start_col is None or end_col is None:
                raise ValueError(f"Invalid column range: {part}")
            for col in range(start_col, end_col + 1):
                columns.add(col)
        else:
            # Single column
            col = range_boundaries(f"{part}1")[0]
            if col is None:
                raise ValueError(f"Invalid column: {part}")
            columns.add(col)

    return sorted(columns)


def _calculate_auto_width(ws, col_idx: int) -> float:
    """Calculate auto-fit width for a column.

    Uses max content length + padding, capped at AUTO_FIT_CAP.
    """
    max_length = 0
    get_column_letter(col_idx)

    # Iterate through all cells in column
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            try:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except Exception:
                pass

    # Calculate width with padding, cap at limit
    if max_length == 0:
        return DEFAULT_WIDTH
    width = min(max_length + AUTO_FIT_PADDING, AUTO_FIT_CAP)
    return float(width)


def _run() -> dict[str, object]:
    parser = create_parser("Set column widths (fixed or auto-fit).")
    add_common_args(parser)
    parser.add_argument(
        "--columns",
        type=str,
        required=True,
        help='Columns to adjust (e.g., "A", "A,C,E", "A:C")',
    )
    parser.add_argument(
        "--width",
        type=str,
        required=True,
        help='Width value (number for fixed width, "auto" for auto-fit)',
    )
    args = parser.parse_args()

    input_path = validate_input_path(args.input)
    output_path = validate_output_path(args.output or str(input_path), create_parents=True)

    file_hash = compute_file_hash(input_path)

    # Parse columns
    try:
        columns = _parse_columns(args.columns)
    except ValueError as e:
        return build_response("error", None, exit_code=1, warnings=[str(e)])

    # Parse width
    is_auto = args.width.lower() == "auto"
    if not is_auto:
        try:
            width_value = float(args.width)
            if width_value < MIN_WIDTH or width_value > MAX_WIDTH:
                return build_response(
                    "error",
                    None,
                    exit_code=1,
                    warnings=[
                        f"Width must be between {MIN_WIDTH} and {MAX_WIDTH}, got {width_value}"
                    ],
                )
        except ValueError:
            return build_response(
                "error",
                None,
                exit_code=1,
                warnings=[f"Invalid width value: {args.width}. Use number or 'auto'"],
            )

    # Load workbook
    wb = load_workbook(str(input_path))
    ws = wb[args.sheet] if args.sheet else wb.active
    if ws is None:
        return build_response("error", None, exit_code=1, warnings=["No active sheet found"])

    # Apply widths
    columns_affected = []
    warnings = []

    for col_idx in columns:
        col_letter = get_column_letter(col_idx)

        if is_auto:
            width = _calculate_auto_width(ws, col_idx)
            if width == DEFAULT_WIDTH:
                warnings.append(f"Column {col_letter} is empty, using default width")
        else:
            width = float(args.width)
            if width > 50:
                warnings.append(
                    f"Column {col_letter} width {width} exceeds 50, may impact readability"
                )

        ws.column_dimensions[col_letter].width = width
        columns_affected.append({"column": col_letter, "width": width})

    # Save workbook
    wb.save(str(output_path))

    # Log to audit trail
    audit = AuditTrail()
    audit.log(
        tool="xls_set_column_width",
        scope="structure:modify",
        target_file=input_path,
        file_version_hash=file_hash,
        actor_nonce="auto",
        operation_details={
            "columns": args.columns,
            "width": args.width,
            "sheet": ws.title,
            "columns_count": len(columns_affected),
        },
        impact={"columns_adjusted": len(columns_affected)},
        success=True,
        exit_code=0,
    )

    return build_response(
        "success",
        {
            "columns": args.columns,
            "width": args.width,
            "sheet": ws.title,
            "columns_affected": columns_affected,
            "is_auto": is_auto,
        },
        warnings=warnings if warnings else None,
    )


def main() -> None:
    run_tool(_run)


if __name__ == "__main__":
    main()
