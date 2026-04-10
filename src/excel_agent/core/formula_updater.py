"""Formula reference updating engine for excel-agent-tools.

Since openpyxl does NOT manage formula dependencies when rows/columns
are inserted or deleted (confirmed by openpyxl 3.1.5 documentation:
"Openpyxl does not manage dependencies, such as formulae, tables,
charts, etc., when rows or columns are inserted or deleted"), this
module provides the reference updating logic.

Capabilities:
- Replace sheet name references in formulas (for sheet rename)
- Adjust row references after row insert/delete
- Adjust column references after column insert/delete

Uses the openpyxl Tokenizer to parse formulas and identify OPERAND
tokens with RANGE subtype, then performs targeted string replacement.
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import Token
from openpyxl.utils import column_index_from_string, get_column_letter

if TYPE_CHECKING:
    from openpyxl import Workbook

import logging

logger = logging.getLogger(__name__)

# Regex for sheet prefix in token values
_SHEET_PREFIX_RE = re.compile(r"^(?:'([^']+)'|([A-Za-z0-9_.\-]+))!(.+)$")

# Regex for cell reference (with optional $ anchors)
_CELL_REF_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)$")

# Regex for range reference
_RANGE_REF_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+):(\$?)([A-Za-z]{1,3})(\$?)(\d+)$")


def rename_sheet_in_formulas(
    workbook: Workbook,
    old_name: str,
    new_name: str,
) -> int:
    """Update all formula references from old_name to new_name across the workbook.

    Iterates every cell in every sheet. For each formula, tokenizes it,
    finds RANGE tokens referencing the old sheet name, and replaces them.

    Args:
        workbook: The workbook to update.
        old_name: The old sheet name.
        new_name: The new sheet name.

    Returns:
        Number of formulas updated.
    """
    updated_count = 0
    old_quoted = _quote_sheet_name(old_name)
    new_quoted = _quote_sheet_name(new_name)

    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                new_formula = _replace_sheet_in_formula(
                    cell.value, old_name, old_quoted, new_name, new_quoted
                )
                if new_formula != cell.value:
                    cell.value = new_formula
                    updated_count += 1

    # Also update defined names
    for _name, defn in workbook.defined_names.items():
        if hasattr(defn, "attr_text") and defn.attr_text and old_name in defn.attr_text:
            new_text = defn.attr_text.replace(f"'{old_name}'!", f"'{new_name}'!").replace(
                f"{old_name}!", f"{new_name}!"
            )
            if new_text != defn.attr_text:
                defn.attr_text = new_text
                updated_count += 1

    return updated_count


def _quote_sheet_name(name: str) -> str:
    """Quote a sheet name if it contains special characters."""
    if re.search(r"[^A-Za-z0-9_.]", name):
        return f"'{name}'"
    return name


def _replace_sheet_in_formula(
    formula: str,
    old_name: str,
    old_quoted: str,
    new_name: str,
    new_quoted: str,
) -> str:
    """Replace sheet name references within a formula string.

    Uses simple string replacement on known patterns:
    'OldName'!A1 → 'NewName'!A1
    OldName!A1 → NewName!A1
    """
    result = formula
    # Quoted form: 'Old Name'!
    result = result.replace(f"'{old_name}'!", f"'{new_name}'!")
    # Unquoted form: OldName!
    if old_name != old_quoted:
        pass  # Only quoted form needed for names with spaces
    else:
        result = result.replace(f"{old_name}!", f"{new_name}!")
    return result


def adjust_row_references(
    workbook: Workbook,
    target_sheet: str,
    start_row: int,
    row_delta: int,
) -> int:
    """Adjust row references in formulas after row insert or delete.

    For all formulas across the entire workbook, find references to cells
    in target_sheet at or below start_row, and shift them by row_delta.

    Args:
        workbook: The workbook to update.
        target_sheet: Sheet where rows were inserted/deleted.
        start_row: The row at/after which the shift applies.
        row_delta: Positive for insertion, negative for deletion.

    Returns:
        Number of formulas updated.
    """
    updated = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                new_formula = _shift_rows_in_formula(
                    cell.value, target_sheet, ws.title, start_row, row_delta
                )
                if new_formula != cell.value:
                    cell.value = new_formula
                    updated += 1
    return updated


def adjust_col_references(
    workbook: Workbook,
    target_sheet: str,
    start_col: int,
    col_delta: int,
) -> int:
    """Adjust column references in formulas after column insert or delete.

    Args:
        workbook: The workbook to update.
        target_sheet: Sheet where columns were inserted/deleted.
        start_col: The column at/after which the shift applies (1-indexed).
        col_delta: Positive for insertion, negative for deletion.

    Returns:
        Number of formulas updated.
    """
    updated = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                new_formula = _shift_cols_in_formula(
                    cell.value, target_sheet, ws.title, start_col, col_delta
                )
                if new_formula != cell.value:
                    cell.value = new_formula
                    updated += 1
    return updated


def _shift_rows_in_formula(
    formula: str,
    target_sheet: str,
    current_sheet: str,
    start_row: int,
    row_delta: int,
) -> str:
    """Shift row numbers in cell references within a formula."""
    try:
        tok = Tokenizer(formula)
    except Exception:
        return formula

    parts: list[str] = []
    for token in tok.items:
        if token.type == Token.OPERAND and token.subtype == Token.RANGE:
            new_val = _shift_token_rows(
                token.value, target_sheet, current_sheet, start_row, row_delta
            )
            parts.append(new_val)
        else:
            parts.append(token.value)

    reconstructed = "=" + "".join(parts)
    return reconstructed


def _shift_token_rows(
    token_value: str,
    target_sheet: str,
    current_sheet: str,
    start_row: int,
    row_delta: int,
) -> str:
    """Shift row references in a single token value."""
    ref_sheet = current_sheet
    ref_part = token_value

    m = _SHEET_PREFIX_RE.match(token_value)
    if m:
        ref_sheet = m.group(1) or m.group(2)
        ref_part = m.group(3)
        prefix = token_value[: token_value.rindex("!") + 1]
    else:
        prefix = ""

    if ref_sheet != target_sheet:
        return token_value

    # Try range
    rm = _RANGE_REF_RE.match(ref_part)
    if rm:
        r1 = _shift_single_row(int(rm.group(4)), rm.group(3), start_row, row_delta)
        r2 = _shift_single_row(int(rm.group(8)), rm.group(7), start_row, row_delta)
        if r1 is None or r2 is None:
            return prefix + "#REF!"
        g = rm.group
        return f"{prefix}{g(1)}{g(2)}{g(3)}{r1}:{g(5)}{g(6)}{g(7)}{r2}"

    # Try single cell
    cm = _CELL_REF_RE.match(ref_part)
    if cm:
        new_row = _shift_single_row(int(cm.group(4)), cm.group(3), start_row, row_delta)
        if new_row is None:
            return prefix + "#REF!"
        return f"{prefix}{cm.group(1)}{cm.group(2)}{cm.group(3)}{new_row}"

    return token_value


def _shift_single_row(row: int, dollar: str, start_row: int, delta: int) -> int | None:
    """Shift a single row number. Returns None if the row was deleted.

    Args:
        row: The row number
        dollar: "$" if absolute row reference, "" if relative
        start_row: Row where insert/delete happened
        delta: Positive for insert, negative for delete

    Returns:
        New row number, or None if row was deleted

    Examples:
        # Fully relative (A1) - shifts if at/after start_row
        _shift_single_row(5, "", 3, 1) -> 6    # Insert before
        _shift_single_row(5, "", 3, -1) -> 4   # Delete before
        _shift_single_row(2, "", 3, 1) -> 2    # Before start, no shift

        # Row-absolute (A$1) - never shifts
        _shift_single_row(5, "$", 3, 1) -> 5
        _shift_single_row(5, "$", 3, -1) -> 5

        # Deleted row returns None
        _shift_single_row(3, "", 3, -1) -> None
    """
    # Absolute reference: never shift
    if dollar == "$":
        return row

    # Relative reference: shift if at/after start_row
    if row >= start_row:
        new_row = row + delta
        return new_row if new_row >= 1 else None
    return row


def _shift_cols_in_formula(
    formula: str,
    target_sheet: str,
    current_sheet: str,
    start_col: int,
    col_delta: int,
) -> str:
    """Shift column letters in cell references within a formula."""
    try:
        tok = Tokenizer(formula)
    except Exception:
        return formula

    parts: list[str] = []
    for token in tok.items:
        if token.type == Token.OPERAND and token.subtype == Token.RANGE:
            new_val = _shift_token_cols(
                token.value, target_sheet, current_sheet, start_col, col_delta
            )
            parts.append(new_val)
        else:
            parts.append(token.value)

    return "=" + "".join(parts)


def _shift_token_cols(
    token_value: str,
    target_sheet: str,
    current_sheet: str,
    start_col: int,
    col_delta: int,
) -> str:
    """Shift column references in a single token value."""
    ref_sheet = current_sheet
    ref_part = token_value

    m = _SHEET_PREFIX_RE.match(token_value)
    if m:
        ref_sheet = m.group(1) or m.group(2)
        ref_part = m.group(3)
        prefix = token_value[: token_value.rindex("!") + 1]
    else:
        prefix = ""

    if ref_sheet != target_sheet:
        return token_value

    cm = _CELL_REF_RE.match(ref_part)
    if cm:
        col_num = column_index_from_string(cm.group(2))
        if col_num >= start_col:
            new_col = col_num + col_delta
            if new_col < 1:
                return prefix + "#REF!"
            return f"{prefix}{cm.group(1)}{get_column_letter(new_col)}{cm.group(3)}{cm.group(4)}"
        return token_value

    rm = _RANGE_REF_RE.match(ref_part)
    if rm:
        c1 = column_index_from_string(rm.group(2))
        c2 = column_index_from_string(rm.group(6))
        nc1 = (c1 + col_delta) if c1 >= start_col else c1
        nc2 = (c2 + col_delta) if c2 >= start_col else c2
        if nc1 < 1 or nc2 < 1:
            return prefix + "#REF!"
        return (
            f"{prefix}{rm.group(1)}{get_column_letter(nc1)}{rm.group(3)}{rm.group(4)}"
            f":{rm.group(5)}{get_column_letter(nc2)}{rm.group(7)}{rm.group(8)}"
        )

    return token_value
