"""
Formula error scanner for excel-agent-tools.

Scans all cells in a workbook for Excel error values:
#REF!, #VALUE!, #DIV/0!, #NAME?, #N/A, #NUM!, #NULL!

These can appear either as cached values in data_only mode or as
literal error strings in formula-preserving mode. We detect both.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_ERROR_VALUES = frozenset(
    {
        "#REF!",
        "#VALUE!",
        "#DIV/0!",
        "#NAME?",
        "#N/A",
        "#NUM!",
        "#NULL!",
    }
)


def detect_errors(workbook: Workbook) -> list[dict[str, Any]]:
    """Scan all cells for formula error values.

    Returns:
        List of dicts: [{"sheet", "cell", "error", "formula"}, ...]
    """
    errors: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                val_str = str(val)
                if val_str in _ERROR_VALUES:
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "error": val_str,
                            "formula": val_str if cell.data_type == "f" else None,
                        }
                    )

    return errors
